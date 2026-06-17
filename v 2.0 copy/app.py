import os
import secrets
import io
import shutil
import threading
import time as _time
from datetime import datetime, date, timedelta
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import Workbook

from database import get_db, init_db, DB_PATH

def _hash_pw(password):
    return generate_password_hash(password, method='pbkdf2:sha256')

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
BACKUP_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backups')
os.makedirs(BACKUP_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}


@app.route('/favicon.ico')
def favicon():
    return send_file(
        os.path.join(app.static_folder, 'favicon.png'),
        mimetype='image/png',
        max_age=0
    )


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def log_activity(user_id, action, details=''):
    db = get_db()
    db.execute('INSERT INTO activity_logs (user_id, action, ip_address, details) VALUES (?, ?, ?, ?)',
               (user_id, action, request.remote_addr if request else '', details))
    db.commit()
    db.close()


def send_notification(user_id, title, message, notif_type='info', link=None):
    db = get_db()
    db.execute('INSERT INTO notifications (user_id, title, message, type, link) VALUES (?, ?, ?, ?, ?)',
               (user_id, title, message, notif_type, link))
    db.commit()
    db.close()

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'


# ── User Model ───────────────────────────────────────────────
class User(UserMixin):
    def __init__(self, id, username, password_hash, full_name, email, phone, pin_hash, profile_pic, role, is_active, created_at):
        self.id = id
        self.username = username
        self.password_hash = password_hash
        self.full_name = full_name
        self.email = email
        self.phone = phone
        self.pin_hash = pin_hash
        self.profile_pic = profile_pic
        self.role = role
        self._is_active = is_active
        self.created_at = created_at

    @property
    def is_active(self):
        return bool(self._is_active)

    def is_super_admin(self):
        return self.role == 'super_admin'

    def is_admin(self):
        return self.role in ('super_admin', 'admin')

    def has_pin(self):
        return bool(self.pin_hash)


@login_manager.user_loader
def load_user(user_id):
    db = get_db()
    row = db.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    db.close()
    if row:
        return User(**dict(row))
    return None


def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_admin():
            flash('Access denied. Admin privileges required.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


def super_admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_super_admin():
            flash('Access denied. Super Admin privileges required.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


@app.context_processor
def inject_notification_count():
    if current_user.is_authenticated:
        db = get_db()
        count = db.execute('SELECT COUNT(*) FROM notifications WHERE user_id = ? AND is_read = 0',
                          (current_user.id,)).fetchone()[0]
        msg_count = db.execute('SELECT COUNT(*) FROM messages WHERE receiver_id = ? AND is_read = 0',
                              (current_user.id,)).fetchone()[0]
        db.close()
        return {'unread_notif_count': count, 'unread_msg_count': msg_count}
    return {'unread_notif_count': 0, 'unread_msg_count': 0}


# ── Auth Routes ──────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        db = get_db()
        row = db.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        db.close()

        if row and check_password_hash(row['password_hash'], password):
            user = User(**dict(row))
            if not user.is_active:
                flash('Your account has been deactivated. Contact admin.', 'danger')
                return render_template('login.html')
            login_user(user)
            # Log the login
            log_activity(user.id, 'login', 'User logged in')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('dashboard'))
        flash('Invalid username or password.', 'danger')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    log_activity(current_user.id, 'logout', 'User logged out')
    logout_user()
    flash('Logged out successfully.', 'success')
    return redirect(url_for('login'))


# ── PIN Management ───────────────────────────────────────────
@app.route('/pin/setup', methods=['GET', 'POST'])
@login_required
def setup_pin():
    if request.method == 'POST':
        pin = request.form.get('pin', '').strip()
        confirm_pin = request.form.get('confirm_pin', '').strip()

        if not pin or len(pin) < 4 or len(pin) > 6 or not pin.isdigit():
            flash('PIN must be 4-6 digits.', 'danger')
        elif pin != confirm_pin:
            flash('PINs do not match.', 'danger')
        else:
            db = get_db()
            db.execute('UPDATE users SET pin_hash = ? WHERE id = ?',
                      (_hash_pw(pin), current_user.id))
            db.commit()
            db.close()
            log_activity(current_user.id, 'pin_setup', 'User set up PIN')
            flash('PIN set successfully.', 'success')
            return redirect(url_for('dashboard'))

    return render_template('pin_setup.html')


@app.route('/pin/verify', methods=['POST'])
@login_required
def verify_pin():
    pin = request.form.get('pin', '').strip()
    db = get_db()
    row = db.execute('SELECT pin_hash FROM users WHERE id = ?', (current_user.id,)).fetchone()
    db.close()

    if not row or not row['pin_hash']:
        return jsonify({'valid': False, 'error': 'PIN not set. Please set up your PIN first.'})

    if check_password_hash(row['pin_hash'], pin):
        return jsonify({'valid': True})
    else:
        return jsonify({'valid': False, 'error': 'Invalid PIN. Please try again.'})


# ── Notifications ────────────────────────────────────────────
@app.route('/notifications')
@login_required
def notifications():
    db = get_db()
    notifs = db.execute('''
        SELECT * FROM notifications WHERE user_id = ?
        ORDER BY created_at DESC LIMIT 100
    ''', (current_user.id,)).fetchall()
    # Mark all as read
    db.execute('UPDATE notifications SET is_read = 1 WHERE user_id = ? AND is_read = 0',
               (current_user.id,))
    db.commit()
    db.close()
    return render_template('notifications.html', notifications=notifs)


@app.route('/notifications/poll')
@login_required
def notifications_poll():
    after_id = request.args.get('after', 0, type=int)
    db = get_db()
    count = db.execute('SELECT COUNT(*) FROM notifications WHERE user_id = ? AND is_read = 0',
                       (current_user.id,)).fetchone()[0]
    msg_count = db.execute('SELECT COUNT(*) FROM messages WHERE receiver_id = ? AND is_read = 0',
                           (current_user.id,)).fetchone()[0]
    new_notifs = []
    if after_id:
        rows = db.execute('''
            SELECT id, title, message, type, link, created_at FROM notifications
            WHERE user_id = ? AND id > ? ORDER BY id ASC LIMIT 10
        ''', (current_user.id, after_id)).fetchall()
        new_notifs = [dict(r) for r in rows]
    else:
        latest = db.execute('SELECT MAX(id) FROM notifications WHERE user_id = ?',
                            (current_user.id,)).fetchone()[0]
        new_notifs = [{'id': latest or 0}]
    db.close()
    return jsonify({'notif_count': count, 'msg_count': msg_count, 'notifications': new_notifs})


@app.route('/tasks/remind/<int:task_id>')
@login_required
def send_task_reminder(task_id):
    db = get_db()
    task = db.execute('SELECT * FROM tasks WHERE id = ?', (task_id,)).fetchone()
    db.close()

    if not task:
        flash('Task not found.', 'danger')
        return redirect(url_for('tasks'))

    if not task['assigned_to']:
        flash('This task is not assigned to anyone.', 'warning')
        return redirect(url_for('tasks'))

    if task['status'] == 'completed':
        flash('This task is already completed.', 'info')
        return redirect(url_for('tasks'))

    send_notification(
        task['assigned_to'],
        'Urgent Task Reminder',
        f'{current_user.full_name} sent you a reminder: Please complete the task "{task["title"]}" — it\'s urgent!',
        'reminder',
        url_for('tasks')
    )
    log_activity(current_user.id, 'task_reminder', f'Sent reminder for task #{task_id}: {task["title"]}')
    flash(f'Reminder sent for task "{task["title"]}".', 'success')
    return redirect(url_for('tasks'))


@app.route('/tasks/accept/<int:task_id>')
@login_required
def accept_task(task_id):
    db = get_db()
    task = db.execute('SELECT * FROM tasks WHERE id = ?', (task_id,)).fetchone()

    if not task or task['assigned_to'] != current_user.id:
        db.close()
        flash('Task not found or not assigned to you.', 'danger')
        return redirect(url_for('tasks'))

    if task['status'] != 'pending':
        db.close()
        flash('Task has already been accepted.', 'info')
        return redirect(url_for('tasks'))

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute('UPDATE tasks SET status = ?, accepted_at = ? WHERE id = ?',
               ('accepted', now, task_id))
    db.commit()
    db.close()

    # Notify the task creator
    send_notification(
        task['created_by'],
        'Task Accepted',
        f'{current_user.full_name} accepted the task: "{task["title"]}"',
        'task_accepted',
        url_for('tasks')
    )
    log_activity(current_user.id, 'accept_task', f'Accepted task #{task_id}: {task["title"]}')
    flash(f'Task "{task["title"]}" accepted.', 'success')
    return redirect(url_for('tasks'))


@app.route('/tasks/start/<int:task_id>')
@login_required
def start_task(task_id):
    db = get_db()
    task = db.execute('SELECT * FROM tasks WHERE id = ?', (task_id,)).fetchone()

    if not task or task['assigned_to'] != current_user.id:
        db.close()
        flash('Task not found or not assigned to you.', 'danger')
        return redirect(url_for('tasks'))

    if task['status'] not in ('pending', 'accepted'):
        db.close()
        flash('Task cannot be started from its current status.', 'info')
        return redirect(url_for('tasks'))

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    accepted_at = task['accepted_at'] or now
    db.execute('UPDATE tasks SET status = ?, accepted_at = ?, started_at = ? WHERE id = ?',
               ('in_progress', accepted_at, now, task_id))
    db.commit()
    db.close()

    # Notify the task creator
    send_notification(
        task['created_by'],
        'Task Started',
        f'{current_user.full_name} started working on the task: "{task["title"]}"',
        'task_started',
        url_for('tasks')
    )
    log_activity(current_user.id, 'start_task', f'Started task #{task_id}: {task["title"]}')
    flash(f'Task "{task["title"]}" started.', 'success')
    return redirect(url_for('tasks'))


@app.route('/tasks/complete/<int:task_id>')
@login_required
def complete_task(task_id):
    db = get_db()
    task = db.execute('SELECT * FROM tasks WHERE id = ?', (task_id,)).fetchone()

    if not task:
        db.close()
        flash('Task not found.', 'danger')
        return redirect(url_for('tasks'))

    if not current_user.is_admin() and task['assigned_to'] != current_user.id:
        db.close()
        flash('You can only complete tasks assigned to you.', 'danger')
        return redirect(url_for('tasks'))

    if task['status'] == 'completed':
        db.close()
        flash('Task is already completed.', 'info')
        return redirect(url_for('tasks'))

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    accepted_at = task['accepted_at'] or now
    started_at = task['started_at'] or now
    db.execute('UPDATE tasks SET status = ?, accepted_at = ?, started_at = ?, completed_at = ? WHERE id = ?',
               ('completed', accepted_at, started_at, now, task_id))
    db.commit()
    db.close()

    # Notify the task creator
    if task['created_by'] and task['created_by'] != current_user.id:
        send_notification(
            task['created_by'],
            'Task Completed',
            f'{current_user.full_name} completed the task: "{task["title"]}"',
            'task_completed',
            url_for('tasks')
        )
    log_activity(current_user.id, 'complete_task', f'Completed task #{task_id}: {task["title"]}')
    flash(f'Task "{task["title"]}" marked as completed.', 'success')
    return redirect(url_for('tasks'))


@app.route('/deliveries/received/<int:delivery_id>', methods=['GET', 'POST'])
@login_required
def receive_delivery(delivery_id):
    db = get_db()
    delivery = db.execute('SELECT * FROM deliveries WHERE id = ?', (delivery_id,)).fetchone()

    if not delivery:
        db.close()
        flash('Delivery not found.', 'danger')
        return redirect(url_for('deliveries'))

    if delivery['signed_note_status'] == 'received':
        db.close()
        flash('Signed note already marked as received.', 'info')
        return redirect(url_for('deliveries'))

    received_by = request.form.get('received_by', current_user.full_name) if request.method == 'POST' else current_user.full_name
    note_copy_type = request.form.get('note_copy_type', 'hard_copy') if request.method == 'POST' else 'hard_copy'

    db.execute('UPDATE deliveries SET signed_note_status = ?, received_by = ?, note_copy_type = ? WHERE id = ?',
               ('received', received_by, note_copy_type, delivery_id))
    db.commit()
    log_activity(current_user.id, 'receive_delivery', f'Marked signed note received for delivery #{delivery_id}: {delivery["delivery_note_number"]} (by {received_by}, {note_copy_type})')
    db.close()
    flash(f'Delivery "{delivery["delivery_note_number"]}" signed note marked as received by {received_by} ({note_copy_type.replace("_", " ").title()}).', 'success')
    return redirect(url_for('deliveries'))


@app.route('/deliveries/pay/<int:delivery_id>', methods=['POST'])
@login_required
def pay_delivery(delivery_id):
    db = get_db()
    delivery = db.execute('SELECT * FROM deliveries WHERE id = ?', (delivery_id,)).fetchone()

    if not delivery:
        db.close()
        flash('Delivery not found.', 'danger')
        return redirect(url_for('deliveries'))

    amount_paid = float(request.form.get('amount_paid') or 0)
    payment_method = request.form.get('payment_method', '')
    paid_date = request.form.get('paid_date') or datetime.now().strftime('%Y-%m-%d')
    paid_by_employee = request.form.get('paid_by_employee', '').strip()

    if amount_paid <= 0:
        db.close()
        flash('Amount must be greater than 0.', 'danger')
        return redirect(url_for('deliveries'))

    if not payment_method:
        db.close()
        flash('Payment method is required.', 'danger')
        return redirect(url_for('deliveries'))

    total_paid = (delivery['amount_paid'] or 0) + amount_paid
    charge = delivery['transportation_charge'] or 0
    charge_paid = 1 if total_paid >= charge else 0

    db.execute('''UPDATE deliveries SET amount_paid=?, charge_paid=?, payment_method=?,
                  paid_date=?, paid_by=?, updated_at=CURRENT_TIMESTAMP WHERE id=?''',
               (total_paid, charge_paid, payment_method, paid_date,
                paid_by_employee or current_user.full_name, delivery_id))
    # Record payment in history
    db.execute('''INSERT INTO delivery_payments (delivery_id, amount, payment_method, paid_by_employee, paid_date, created_by)
                  VALUES (?, ?, ?, ?, ?, ?)''',
               (delivery_id, amount_paid, payment_method, paid_by_employee or current_user.full_name,
                paid_date, current_user.id))
    db.commit()
    log_activity(current_user.id, 'pay_delivery',
                 f'Payment of {amount_paid} for delivery #{delivery_id}: {delivery["delivery_note_number"]} (paid by: {paid_by_employee or current_user.full_name})')
    db.close()
    if charge_paid:
        flash(f'Delivery "{delivery["delivery_note_number"]}" fully paid.', 'success')
    else:
        flash(f'Payment of {amount_paid:.2f} recorded for "{delivery["delivery_note_number"]}". Balance: {charge - total_paid:.2f}', 'info')
    return redirect(url_for('deliveries'))


# ── Profile Picture ──────────────────────────────────────────
@app.route('/profile/upload-pic/<int:user_id>', methods=['POST'])
@login_required
def upload_profile_pic(user_id):
    # Users can upload their own pic; admins can upload for any user
    if user_id != current_user.id and not current_user.is_super_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))

    if 'profile_pic' not in request.files:
        flash('No file selected.', 'danger')
        return redirect(request.referrer or url_for('dashboard'))

    file = request.files['profile_pic']
    if file.filename == '':
        flash('No file selected.', 'danger')
        return redirect(request.referrer or url_for('dashboard'))

    if file and allowed_file(file.filename):
        ext = secure_filename(file.filename).rsplit('.', 1)[1].lower()
        filename = f'profile_{user_id}.{ext}'
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)

        db = get_db()
        db.execute('UPDATE users SET profile_pic = ? WHERE id = ?', (filename, user_id))
        db.commit()
        db.close()
        log_activity(current_user.id, 'upload_profile_pic', f'Updated profile picture for user #{user_id}')
        flash('Profile picture updated.', 'success')
    else:
        flash('Invalid file type. Use PNG, JPG, GIF or WEBP.', 'danger')

    return redirect(request.referrer or url_for('dashboard'))


# ── Dashboard ────────────────────────────────────────────────
@app.route('/')
@login_required
def dashboard():
    db = get_db()

    # Task stats
    if current_user.is_admin():
        total_tasks = db.execute('SELECT COUNT(*) FROM tasks').fetchone()[0]
        pending_tasks = db.execute("SELECT COUNT(*) FROM tasks WHERE status='pending'").fetchone()[0]
        in_progress_tasks = db.execute("SELECT COUNT(*) FROM tasks WHERE status='in_progress'").fetchone()[0]
        completed_tasks = db.execute("SELECT COUNT(*) FROM tasks WHERE status='completed'").fetchone()[0]
    else:
        total_tasks = db.execute('SELECT COUNT(*) FROM tasks WHERE assigned_to = ?', (current_user.id,)).fetchone()[0]
        pending_tasks = db.execute("SELECT COUNT(*) FROM tasks WHERE assigned_to = ? AND status='pending'", (current_user.id,)).fetchone()[0]
        in_progress_tasks = db.execute("SELECT COUNT(*) FROM tasks WHERE assigned_to = ? AND status='in_progress'", (current_user.id,)).fetchone()[0]
        completed_tasks = db.execute("SELECT COUNT(*) FROM tasks WHERE assigned_to = ? AND status='completed'", (current_user.id,)).fetchone()[0]

    # Delivery stats
    total_deliveries = db.execute('SELECT COUNT(*) FROM deliveries').fetchone()[0]
    pending_payments = db.execute("SELECT COUNT(*) FROM deliveries WHERE charge_paid = 0 AND transportation_charge > 0").fetchone()[0]
    pending_notes = db.execute("SELECT COUNT(*) FROM deliveries WHERE signed_note_status = 'pending'").fetchone()[0]

    # Recent tasks
    if current_user.is_admin():
        recent_tasks = db.execute('''
            SELECT t.*, u.full_name as assigned_name, c.full_name as creator_name
            FROM tasks t
            LEFT JOIN users u ON t.assigned_to = u.id
            LEFT JOIN users c ON t.created_by = c.id
            ORDER BY t.created_at DESC LIMIT 5
        ''').fetchall()
    else:
        recent_tasks = db.execute('''
            SELECT t.*, u.full_name as assigned_name, c.full_name as creator_name
            FROM tasks t
            LEFT JOIN users u ON t.assigned_to = u.id
            LEFT JOIN users c ON t.created_by = c.id
            WHERE t.assigned_to = ?
            ORDER BY t.created_at DESC LIMIT 5
        ''', (current_user.id,)).fetchall()

    # Recent deliveries
    recent_deliveries = db.execute('''
        SELECT d.*, u.full_name as creator_name
        FROM deliveries d
        LEFT JOIN users u ON d.created_by = u.id
        ORDER BY d.created_at DESC LIMIT 5
    ''').fetchall()

    # HR stats
    total_employees = db.execute('SELECT COUNT(*) FROM hr_employees WHERE is_active = 1').fetchone()[0]
    pending_leaves = db.execute("SELECT COUNT(*) FROM hr_leaves WHERE status = 'pending'").fetchone()[0]

    # Vehicle doc expiry stats
    total_vehicles = db.execute('SELECT COUNT(*) FROM vehicles WHERE is_active = 1').fetchone()[0]
    expired_docs = 0
    expiring_docs = 0
    all_vdocs = db.execute('SELECT expiry_date, reminder_days FROM vehicle_documents vd JOIN vehicles v ON vd.vehicle_id = v.id WHERE v.is_active = 1').fetchall()
    today_date = date.today()
    for vd in all_vdocs:
        try:
            exp = datetime.strptime(vd['expiry_date'], '%Y-%m-%d').date()
            if exp < today_date:
                expired_docs += 1
            elif (exp - timedelta(days=vd['reminder_days'])) <= today_date:
                expiring_docs += 1
        except Exception:
            pass

    db.close()

    return render_template('dashboard.html',
        total_tasks=total_tasks, pending_tasks=pending_tasks,
        in_progress_tasks=in_progress_tasks, completed_tasks=completed_tasks,
        total_deliveries=total_deliveries, pending_payments=pending_payments,
        pending_notes=pending_notes, recent_tasks=recent_tasks,
        recent_deliveries=recent_deliveries,
        total_employees=total_employees, pending_leaves=pending_leaves,
        total_vehicles=total_vehicles, expired_docs=expired_docs, expiring_docs=expiring_docs)


# ── Task Management ──────────────────────────────────────────
@app.route('/tasks')
@login_required
def tasks():
    db = get_db()
    status_filter = request.args.get('status', '')
    user_filter = request.args.get('user', '')
    priority_filter = request.args.get('priority', '')
    group_filter = request.args.get('group', '')

    query = '''
        SELECT t.*, u.full_name as assigned_name, u.phone as assigned_phone, c.full_name as creator_name,
               g.name as group_name, g.color as group_color
        FROM tasks t
        LEFT JOIN users u ON t.assigned_to = u.id
        LEFT JOIN users c ON t.created_by = c.id
        LEFT JOIN task_groups g ON t.group_id = g.id
        WHERE 1=1
    '''
    params = []

    if not current_user.is_admin():
        query += ' AND t.assigned_to = ?'
        params.append(current_user.id)

    if status_filter:
        query += ' AND t.status = ?'
        params.append(status_filter)

    if priority_filter:
        query += ' AND t.priority = ?'
        params.append(priority_filter)

    if user_filter and current_user.is_admin():
        query += ' AND t.assigned_to = ?'
        params.append(int(user_filter))

    if group_filter:
        query += ' AND t.group_id = ?'
        params.append(int(group_filter))

    query += ' ORDER BY t.created_at DESC'

    task_list = db.execute(query, params).fetchall()
    users = db.execute("SELECT id, full_name FROM users WHERE is_active = 1").fetchall()
    groups = db.execute("SELECT id, name, color FROM task_groups ORDER BY name").fetchall()
    db.close()

    return render_template('tasks.html', tasks=task_list, users=users, groups=groups,
                           status_filter=status_filter, user_filter=user_filter,
                           priority_filter=priority_filter, group_filter=group_filter)


@app.route('/tasks/add', methods=['GET', 'POST'])
@login_required
def add_task():
    db = get_db()
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        assigned_to = request.form.get('assigned_to')
        priority = request.form.get('priority', 'medium')
        group_id = request.form.get('group_id')

        if not title:
            flash('Task title is required.', 'danger')
        else:
            db.execute('''
                INSERT INTO tasks (title, description, assigned_to, created_by, priority, group_id)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (title, description, int(assigned_to) if assigned_to else None, current_user.id, priority,
                  int(group_id) if group_id else None))
            db.commit()
            log_activity(current_user.id, 'add_task', f'Added task: {title}')

            # Send notification to assigned person
            if assigned_to:
                send_notification(
                    int(assigned_to),
                    'New Task Assigned',
                    f'{current_user.full_name} assigned you a new task: "{title}" (Priority: {priority.title()})',
                    'task_assigned',
                    url_for('tasks')
                )

            db.close()
            flash('Task added successfully.', 'success')
            return redirect(url_for('tasks'))

    users = db.execute("SELECT id, full_name FROM users WHERE is_active = 1").fetchall()
    groups = db.execute("SELECT id, name, color FROM task_groups ORDER BY name").fetchall()
    db.close()
    return render_template('task_form.html', task=None, users=users, groups=groups)


@app.route('/tasks/edit/<int:task_id>', methods=['GET', 'POST'])
@login_required
def edit_task(task_id):
    db = get_db()
    task = db.execute('SELECT * FROM tasks WHERE id = ?', (task_id,)).fetchone()

    if not task:
        db.close()
        flash('Task not found.', 'danger')
        return redirect(url_for('tasks'))

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        assigned_to = request.form.get('assigned_to')
        priority = request.form.get('priority', 'medium')
        status = request.form.get('status', 'pending')
        group_id = request.form.get('group_id')

        completed_at = None
        if status == 'completed' and task['status'] != 'completed':
            completed_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elif status == 'completed' and task['completed_at']:
            completed_at = task['completed_at']

        accepted_at = task['accepted_at']
        started_at = task['started_at']
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if status in ('accepted', 'in_progress', 'completed') and not accepted_at:
            accepted_at = now
        if status in ('in_progress', 'completed') and not started_at:
            started_at = now

        if not title:
            flash('Task title is required.', 'danger')
        else:
            db.execute('''
                UPDATE tasks SET title=?, description=?, assigned_to=?, priority=?, status=?, accepted_at=?, started_at=?, completed_at=?, group_id=?
                WHERE id=?
            ''', (title, description, int(assigned_to) if assigned_to else None,
                  priority, status, accepted_at, started_at, completed_at,
                  int(group_id) if group_id else None, task_id))
            db.commit()
            log_activity(current_user.id, 'edit_task', f'Edited task #{task_id}: {title} [Status: {status}]')

            # Notify if assigned person changed
            new_assigned = int(assigned_to) if assigned_to else None
            old_assigned = task['assigned_to']
            if new_assigned and new_assigned != old_assigned:
                send_notification(
                    new_assigned,
                    'Task Reassigned to You',
                    f'{current_user.full_name} assigned you a task: "{title}" (Priority: {priority.title()})',
                    'task_assigned',
                    url_for('tasks')
                )

            db.close()
            flash('Task updated successfully.', 'success')
            return redirect(url_for('tasks'))

    users = db.execute("SELECT id, full_name FROM users WHERE is_active = 1").fetchall()
    groups = db.execute("SELECT id, name, color FROM task_groups ORDER BY name").fetchall()
    db.close()
    return render_template('task_form.html', task=task, users=users, groups=groups)


@app.route('/tasks/delete/<int:task_id>')
@admin_required
def delete_task(task_id):
    db = get_db()
    task = db.execute('SELECT title FROM tasks WHERE id = ?', (task_id,)).fetchone()
    db.execute('DELETE FROM tasks WHERE id = ?', (task_id,))
    db.commit()
    log_activity(current_user.id, 'delete_task', f'Deleted task #{task_id}: {task["title"] if task else "Unknown"}')
    db.close()
    flash('Task deleted.', 'success')
    return redirect(url_for('tasks'))


@app.route('/tasks/export')
@login_required
def export_tasks():
    db = get_db()
    status_filter = request.args.get('status', '')
    user_filter = request.args.get('user', '')
    priority_filter = request.args.get('priority', '')

    query = '''
        SELECT t.*, u.full_name as assigned_name, c.full_name as creator_name, g.name as group_name
        FROM tasks t
        LEFT JOIN users u ON t.assigned_to = u.id
        LEFT JOIN users c ON t.created_by = c.id
        LEFT JOIN task_groups g ON t.group_id = g.id
        WHERE 1=1
    '''
    params = []
    if not current_user.is_admin():
        query += ' AND t.assigned_to = ?'
        params.append(current_user.id)
    if status_filter:
        query += ' AND t.status = ?'
        params.append(status_filter)
    if priority_filter:
        query += ' AND t.priority = ?'
        params.append(priority_filter)
    if user_filter and current_user.is_admin():
        query += ' AND t.assigned_to = ?'
        params.append(int(user_filter))
    query += ' ORDER BY t.created_at DESC'

    task_list = db.execute(query, params).fetchall()
    db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tasks'
    ws.append(['#', 'Title', 'Group', 'Description', 'Assigned To', 'Created By', 'Priority', 'Status',
               'Created', 'Accepted', 'Started', 'Completed'])
    for t in task_list:
        ws.append([t['id'], t['title'], t['group_name'] or '', t['description'] or '', t['assigned_name'] or 'Unassigned',
                   t['creator_name'], t['priority'], t['status'],
                   t['created_at'] or '', t['accepted_at'] or '', t['started_at'] or '', t['completed_at'] or ''])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='tasks.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Task Groups ──────────────────────────────────────────────
@app.route('/task-groups')
@admin_required
def task_groups():
    db = get_db()
    groups = db.execute('''
        SELECT g.*, u.full_name as creator_name,
               (SELECT COUNT(*) FROM tasks WHERE group_id = g.id) as task_count
        FROM task_groups g
        LEFT JOIN users u ON g.created_by = u.id
        ORDER BY g.name
    ''').fetchall()
    db.close()
    return render_template('task_groups.html', groups=groups)


@app.route('/task-groups/add', methods=['POST'])
@admin_required
def add_task_group():
    name = request.form.get('name', '').strip()
    color = request.form.get('color', '#6c757d').strip()
    if not name:
        flash('Group name is required.', 'danger')
        return redirect(url_for('task_groups'))
    db = get_db()
    existing = db.execute('SELECT id FROM task_groups WHERE name = ?', (name,)).fetchone()
    if existing:
        flash('A group with that name already exists.', 'warning')
        db.close()
        return redirect(url_for('task_groups'))
    db.execute('INSERT INTO task_groups (name, color, created_by) VALUES (?, ?, ?)',
               (name, color, current_user.id))
    db.commit()
    log_activity(current_user.id, 'add_task_group', f'Created task group: {name}')
    db.close()
    flash(f'Group "{name}" created.', 'success')
    return redirect(url_for('task_groups'))


@app.route('/task-groups/edit/<int:group_id>', methods=['POST'])
@admin_required
def edit_task_group(group_id):
    name = request.form.get('name', '').strip()
    color = request.form.get('color', '#6c757d').strip()
    if not name:
        flash('Group name is required.', 'danger')
        return redirect(url_for('task_groups'))
    db = get_db()
    existing = db.execute('SELECT id FROM task_groups WHERE name = ? AND id != ?', (name, group_id)).fetchone()
    if existing:
        flash('A group with that name already exists.', 'warning')
        db.close()
        return redirect(url_for('task_groups'))
    db.execute('UPDATE task_groups SET name = ?, color = ? WHERE id = ?', (name, color, group_id))
    db.commit()
    log_activity(current_user.id, 'edit_task_group', f'Updated task group #{group_id}: {name}')
    db.close()
    flash(f'Group "{name}" updated.', 'success')
    return redirect(url_for('task_groups'))


@app.route('/task-groups/delete/<int:group_id>')
@admin_required
def delete_task_group(group_id):
    db = get_db()
    group = db.execute('SELECT name FROM task_groups WHERE id = ?', (group_id,)).fetchone()
    db.execute('DELETE FROM task_groups WHERE id = ?', (group_id,))
    db.commit()
    log_activity(current_user.id, 'delete_task_group', f'Deleted task group #{group_id}: {group["name"] if group else "Unknown"}')
    db.close()
    flash('Task group deleted.', 'success')
    return redirect(url_for('task_groups'))


# ── Delivery Tracking ────────────────────────────────────────
@app.route('/deliveries')
@login_required
def deliveries():
    db = get_db()
    payment_filter = request.args.get('payment', '')
    note_filter = request.args.get('note_status', '')
    dn_search = request.args.get('dn_search', '').strip()
    driver_filter = request.args.get('driver', '')

    query = '''
        SELECT d.*, u.full_name as creator_name, dp.name as driver_name, dp.vehicle_no as driver_vehicle, dp.mobile as driver_mobile
        FROM deliveries d
        LEFT JOIN users u ON d.created_by = u.id
        LEFT JOIN delivery_persons dp ON d.delivery_person_id = dp.id
        WHERE 1=1
    '''
    params = []

    if dn_search:
        query += ' AND d.delivery_note_number LIKE ?'
        params.append(f'%{dn_search}%')

    if driver_filter:
        query += ' AND d.delivery_person_id = ?'
        params.append(int(driver_filter))

    if payment_filter == 'paid':
        query += ' AND d.charge_paid = 1'
    elif payment_filter == 'unpaid':
        query += ' AND d.charge_paid = 0 AND d.transportation_charge > 0'

    if note_filter:
        query += ' AND d.signed_note_status = ?'
        params.append(note_filter)

    query += ' ORDER BY d.created_at DESC'

    delivery_list = db.execute(query, params).fetchall()
    drivers = db.execute("SELECT id, name FROM delivery_persons WHERE is_active = 1 ORDER BY name").fetchall()
    users_list = db.execute("SELECT id, full_name FROM users WHERE is_active = 1 ORDER BY full_name").fetchall()
    db.close()

    return render_template('deliveries.html', deliveries=delivery_list,
                           payment_filter=payment_filter, note_filter=note_filter,
                           dn_search=dn_search, driver_filter=driver_filter, drivers=drivers,
                           users=users_list)


@app.route('/api/customers')
@login_required
def api_customers():
    """AJAX endpoint for customer name autocomplete."""
    db = get_db()
    customers = db.execute('SELECT name FROM customers ORDER BY name').fetchall()
    db.close()
    return jsonify([c['name'] for c in customers])


@app.route('/deliveries/add', methods=['GET', 'POST'])
@login_required
def add_delivery():
    if request.method == 'POST':
        delivery_note_number = request.form.get('delivery_note_number', '').strip()
        description = request.form.get('description', '').strip()
        customer_name = request.form.get('customer_name', '').strip()
        location_from = request.form.get('location_from', '').strip()
        location_to = request.form.get('location_to', '').strip()
        delivery_date = request.form.get('delivery_date') or None
        delivery_person_id = request.form.get('delivery_person_id') or None
        transportation_charge = float(request.form.get('transportation_charge') or 0)
        charge_paid = 1 if request.form.get('charge_paid') else 0
        amount_paid = float(request.form.get('amount_paid') or 0)
        paid_date = request.form.get('paid_date') or None
        paid_by = request.form.get('paid_by', '').strip()
        payment_method = request.form.get('payment_method', '')
        narration = request.form.get('narration', '').strip()
        signed_note_status = request.form.get('signed_note_status', 'pending')

        if not delivery_note_number or not customer_name:
            flash('Delivery note number and customer name are required.', 'danger')
            db = get_db()
            persons = db.execute("SELECT * FROM delivery_persons WHERE is_active = 1 ORDER BY name").fetchall()
            users_list = db.execute("SELECT id, full_name FROM users WHERE is_active = 1").fetchall()
            db.close()
            return render_template('delivery_form.html', delivery=None, persons=persons, users=users_list)

        db = get_db()
        db.execute('''
            INSERT INTO deliveries (delivery_note_number, description, customer_name,
                location_from, location_to,
                delivery_date, delivery_person_id, transportation_charge, charge_paid, amount_paid, paid_date, paid_by,
                payment_method, narration, signed_note_status, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (delivery_note_number, description, customer_name, location_from, location_to,
              delivery_date,
              int(delivery_person_id) if delivery_person_id else None,
              transportation_charge, charge_paid, amount_paid, paid_date, paid_by,
              payment_method, narration, signed_note_status, current_user.id))
        # Auto-save customer name
        db.execute('INSERT OR IGNORE INTO customers (name) VALUES (?)', (customer_name,))
        db.commit()
        log_activity(current_user.id, 'add_delivery', f'Added delivery: {delivery_note_number} - {customer_name}')
        db.close()
        flash('Delivery added successfully.', 'success')
        return redirect(url_for('deliveries'))

    db = get_db()
    persons = db.execute("SELECT * FROM delivery_persons WHERE is_active = 1 ORDER BY name").fetchall()
    users_list = db.execute("SELECT id, full_name FROM users WHERE is_active = 1").fetchall()
    db.close()
    return render_template('delivery_form.html', delivery=None, persons=persons, users=users_list)


@app.route('/deliveries/edit/<int:delivery_id>', methods=['GET', 'POST'])
@login_required
def edit_delivery(delivery_id):
    db = get_db()
    delivery = db.execute('SELECT * FROM deliveries WHERE id = ?', (delivery_id,)).fetchone()

    if not delivery:
        db.close()
        flash('Delivery not found.', 'danger')
        return redirect(url_for('deliveries'))

    if request.method == 'POST':
        delivery_note_number = request.form.get('delivery_note_number', '').strip()
        description = request.form.get('description', '').strip()
        customer_name = request.form.get('customer_name', '').strip()
        location_from = request.form.get('location_from', '').strip()
        location_to = request.form.get('location_to', '').strip()
        delivery_date = request.form.get('delivery_date') or None
        delivery_person_id = request.form.get('delivery_person_id') or None
        transportation_charge = float(request.form.get('transportation_charge') or 0)
        charge_paid = 1 if request.form.get('charge_paid') else 0
        amount_paid = float(request.form.get('amount_paid') or 0)
        paid_date = request.form.get('paid_date') or None
        paid_by = request.form.get('paid_by', '').strip()
        payment_method = request.form.get('payment_method', '')
        narration = request.form.get('narration', '').strip()
        signed_note_status = request.form.get('signed_note_status', 'pending')

        # Auto-fill payment details when marking as paid
        if charge_paid and not delivery['charge_paid']:
            if not paid_date:
                paid_date = datetime.now().strftime('%Y-%m-%d')
            if not paid_by:
                paid_by = current_user.full_name

        # Require payment details when marking as paid
        if charge_paid and transportation_charge > 0:
            if not payment_method:
                flash('Payment method is required when marking charge as paid.', 'danger')
                persons = db.execute("SELECT * FROM delivery_persons WHERE is_active = 1 ORDER BY name").fetchall()
                users_list = db.execute("SELECT id, full_name FROM users WHERE is_active = 1").fetchall()
                db.close()
                return render_template('delivery_form.html', delivery=delivery, persons=persons, users=users_list)

        if not delivery_note_number or not customer_name:
            flash('Delivery note number and customer name are required.', 'danger')
        else:
            db.execute('''
                UPDATE deliveries SET delivery_note_number=?, description=?, customer_name=?,
                    location_from=?, location_to=?,
                    delivery_date=?, delivery_person_id=?, transportation_charge=?, charge_paid=?, amount_paid=?,
                    paid_date=?, paid_by=?, payment_method=?, narration=?, signed_note_status=?,
                    updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            ''', (delivery_note_number, description, customer_name, location_from, location_to,
                  delivery_date,
                  int(delivery_person_id) if delivery_person_id else None,
                  transportation_charge, charge_paid, amount_paid, paid_date, paid_by,
                  payment_method, narration, signed_note_status, delivery_id))
            # Auto-save customer name
            db.execute('INSERT OR IGNORE INTO customers (name) VALUES (?)', (customer_name,))
            db.commit()
            log_activity(current_user.id, 'edit_delivery', f'Edited delivery #{delivery_id}: {delivery_note_number} - {customer_name}')
            db.close()
            flash('Delivery updated successfully.', 'success')
            return redirect(url_for('deliveries'))

    persons = db.execute("SELECT * FROM delivery_persons WHERE is_active = 1 ORDER BY name").fetchall()
    users_list = db.execute("SELECT id, full_name FROM users WHERE is_active = 1").fetchall()
    db.close()
    return render_template('delivery_form.html', delivery=delivery, persons=persons, users=users_list)


@app.route('/deliveries/delete/<int:delivery_id>')
@admin_required
def delete_delivery(delivery_id):
    db = get_db()
    delivery = db.execute('SELECT delivery_note_number, customer_name FROM deliveries WHERE id = ?', (delivery_id,)).fetchone()
    db.execute('DELETE FROM deliveries WHERE id = ?', (delivery_id,))
    db.commit()
    log_activity(current_user.id, 'delete_delivery', f'Deleted delivery #{delivery_id}: {delivery["delivery_note_number"] + " - " + delivery["customer_name"] if delivery else "Unknown"}')
    db.close()
    flash('Delivery deleted.', 'success')
    return redirect(url_for('deliveries'))


@app.route('/deliveries/export')
@login_required
def export_deliveries():
    db = get_db()
    payment_filter = request.args.get('payment', '')
    note_filter = request.args.get('note_status', '')
    dn_search = request.args.get('dn_search', '').strip()
    driver_filter = request.args.get('driver', '')

    query = '''
        SELECT d.*, u.full_name as creator_name, dp.name as driver_name, dp.vehicle_no as driver_vehicle
        FROM deliveries d
        LEFT JOIN users u ON d.created_by = u.id
        LEFT JOIN delivery_persons dp ON d.delivery_person_id = dp.id
        WHERE 1=1
    '''
    params = []
    if dn_search:
        query += ' AND d.delivery_note_number LIKE ?'
        params.append(f'%{dn_search}%')
    if driver_filter:
        query += ' AND d.delivery_person_id = ?'
        params.append(int(driver_filter))
    if payment_filter == 'paid':
        query += ' AND d.charge_paid = 1'
    elif payment_filter == 'unpaid':
        query += ' AND d.charge_paid = 0 AND d.transportation_charge > 0'
    if note_filter:
        query += ' AND d.signed_note_status = ?'
        params.append(note_filter)
    query += ' ORDER BY d.created_at DESC'

    delivery_list = db.execute(query, params).fetchall()
    db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Deliveries'
    ws.append(['Note #', 'Customer', 'Description', 'Driver', 'Vehicle', 'Delivery Date',
               'Transport Charge', 'Payment Status', 'Payment Method', 'Paid Date', 'Paid By',
               'Signed Note', 'Narration', 'Created By', 'Created'])
    for d in delivery_list:
        ws.append([d['delivery_note_number'], d['customer_name'], d['description'] or '',
                   d['driver_name'] or '', d['driver_vehicle'] or '', d['delivery_date'] or '',
                   d['transportation_charge'] or 0, 'Paid' if d['charge_paid'] else 'Unpaid',
                   d['payment_method'] or '', d['paid_date'] or '', d['paid_by'] or '',
                   d['signed_note_status'], d['narration'] or '', d['creator_name'] or '', d['created_at'] or ''])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='deliveries.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── User Management (Admin) ─────────────────────────────────
@app.route('/users')
@admin_required
def users():
    db = get_db()
    user_list = db.execute('SELECT * FROM users ORDER BY created_at DESC').fetchall()
    db.close()
    return render_template('users.html', users=user_list)


@app.route('/users/add', methods=['GET', 'POST'])
@super_admin_required
def add_user():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        full_name = request.form.get('full_name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        role = request.form.get('role', 'user')

        if not username or not password or not full_name:
            flash('Username, password and full name are required.', 'danger')
            return render_template('user_form.html', user=None)

        db = get_db()
        existing = db.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone()
        if existing:
            db.close()
            flash('Username already exists.', 'danger')
            return render_template('user_form.html', user=None)

        db.execute('''
            INSERT INTO users (username, password_hash, full_name, email, phone, role)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (username, _hash_pw(password), full_name, email, phone, role))
        db.commit()
        log_activity(current_user.id, 'add_user', f'Created user: {username} ({full_name}) - Role: {role}')
        db.close()
        flash('User created successfully.', 'success')
        return redirect(url_for('users'))

    return render_template('user_form.html', user=None)


@app.route('/users/edit/<int:user_id>', methods=['GET', 'POST'])
@super_admin_required
def edit_user(user_id):
    db = get_db()
    user = db.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()

    if not user:
        db.close()
        flash('User not found.', 'danger')
        return redirect(url_for('users'))

    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        role = request.form.get('role', 'user')
        is_active = 1 if request.form.get('is_active') else 0
        new_password = request.form.get('password', '').strip()

        if not full_name:
            flash('Full name is required.', 'danger')
        else:
            if new_password:
                db.execute('''
                    UPDATE users SET full_name=?, email=?, phone=?, role=?, is_active=?, password_hash=?
                    WHERE id=?
                ''', (full_name, email, phone, role, is_active, _hash_pw(new_password), user_id))
            else:
                db.execute('''
                    UPDATE users SET full_name=?, email=?, phone=?, role=?, is_active=?
                    WHERE id=?
                ''', (full_name, email, phone, role, is_active, user_id))
            db.commit()
            log_activity(current_user.id, 'edit_user', f'Edited user #{user_id}: {full_name} - Role: {role}, Active: {is_active}')
            db.close()
            flash('User updated successfully.', 'success')
            return redirect(url_for('users'))

    db.close()
    return render_template('user_form.html', user=user)


@app.route('/users/toggle/<int:user_id>')
@super_admin_required
def toggle_user(user_id):
    if user_id == current_user.id:
        flash('You cannot deactivate yourself.', 'danger')
        return redirect(url_for('users'))

    db = get_db()
    user = db.execute('SELECT is_active FROM users WHERE id = ?', (user_id,)).fetchone()
    if user:
        new_status = 0 if user['is_active'] else 1
        db.execute('UPDATE users SET is_active = ? WHERE id = ?', (new_status, user_id))
        db.commit()
        log_activity(current_user.id, 'toggle_user', f'User #{user_id} {"activated" if new_status else "deactivated"}')
        flash('User status updated.', 'success')
    db.close()
    return redirect(url_for('users'))


# ── Delivery Person Management ────────────────────────────────
@app.route('/delivery-persons')
@login_required
def delivery_persons():
    db = get_db()
    persons = db.execute('SELECT * FROM delivery_persons ORDER BY name').fetchall()
    db.close()
    return render_template('delivery_persons.html', persons=persons)


@app.route('/delivery-persons/quick-add', methods=['POST'])
@login_required
def quick_add_delivery_person():
    """AJAX endpoint to add a driver inline from the delivery form."""
    name = request.form.get('name', '').strip()
    mobile = request.form.get('mobile', '').strip()
    iqama_id = request.form.get('iqama_id', '').strip()
    vehicle_no = request.form.get('vehicle_no', '').strip()
    vehicle_type = request.form.get('vehicle_type', '').strip()
    care_of = request.form.get('care_of', '').strip()

    if not name:
        return jsonify({'error': 'Driver name is required'}), 400

    db = get_db()
    cursor = db.execute(
        'INSERT INTO delivery_persons (name, mobile, iqama_id, vehicle_no, vehicle_type, care_of) VALUES (?, ?, ?, ?, ?, ?)',
        (name, mobile, iqama_id, vehicle_no, vehicle_type, care_of)
    )
    db.commit()
    new_id = cursor.lastrowid
    log_activity(current_user.id, 'add_delivery_person', f'Quick-added driver: {name}')
    db.close()
    return jsonify({'id': new_id, 'name': name, 'mobile': mobile, 'vehicle_no': vehicle_no, 'vehicle_type': vehicle_type})


@app.route('/delivery-persons/add', methods=['GET', 'POST'])
@login_required
def add_delivery_person():
    db = get_db()
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        mobile = request.form.get('mobile', '').strip()
        iqama_id = request.form.get('iqama_id', '').strip()
        vehicle_no = request.form.get('vehicle_no', '').strip()
        vehicle_type = request.form.get('vehicle_type', '').strip()
        care_of = request.form.get('care_of', '').strip()

        if not name:
            flash('Name is required.', 'danger')
            vtypes = db.execute('SELECT name FROM vehicle_types ORDER BY name').fetchall()
            db.close()
            return render_template('delivery_person_form.html', person=None, vehicle_types=vtypes)

        db.execute('''
            INSERT INTO delivery_persons (name, mobile, iqama_id, vehicle_no, vehicle_type, care_of)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (name, mobile, iqama_id, vehicle_no, vehicle_type, care_of))
        db.commit()
        log_activity(current_user.id, 'add_delivery_person', f'Added delivery person: {name} - Mobile: {mobile}, Vehicle: {vehicle_no}')
        db.close()
        flash('Delivery person added successfully.', 'success')
        return redirect(url_for('delivery_persons'))

    vtypes = db.execute('SELECT name FROM vehicle_types ORDER BY name').fetchall()
    db.close()
    return render_template('delivery_person_form.html', person=None, vehicle_types=vtypes)


@app.route('/delivery-persons/edit/<int:person_id>', methods=['GET', 'POST'])
@login_required
def edit_delivery_person(person_id):
    db = get_db()
    person = db.execute('SELECT * FROM delivery_persons WHERE id = ?', (person_id,)).fetchone()

    if not person:
        db.close()
        flash('Delivery person not found.', 'danger')
        return redirect(url_for('delivery_persons'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        mobile = request.form.get('mobile', '').strip()
        iqama_id = request.form.get('iqama_id', '').strip()
        vehicle_no = request.form.get('vehicle_no', '').strip()
        vehicle_type = request.form.get('vehicle_type', '').strip()
        care_of = request.form.get('care_of', '').strip()
        is_active = 1 if request.form.get('is_active') else 0

        if not name:
            flash('Name is required.', 'danger')
        else:
            db.execute('''
                UPDATE delivery_persons SET name=?, mobile=?, iqama_id=?, vehicle_no=?, vehicle_type=?, care_of=?, is_active=?
                WHERE id=?
            ''', (name, mobile, iqama_id, vehicle_no, vehicle_type, care_of, is_active, person_id))
            db.commit()
            log_activity(current_user.id, 'edit_delivery_person', f'Edited delivery person #{person_id}: {name}')
            db.close()
            flash('Delivery person updated successfully.', 'success')
            return redirect(url_for('delivery_persons'))

    vtypes = db.execute('SELECT name FROM vehicle_types ORDER BY name').fetchall()
    db.close()
    return render_template('delivery_person_form.html', person=person, vehicle_types=vtypes)


@app.route('/delivery-persons/delete/<int:person_id>')
@admin_required
def delete_delivery_person(person_id):
    db = get_db()
    person = db.execute('SELECT name FROM delivery_persons WHERE id = ?', (person_id,)).fetchone()
    db.execute('DELETE FROM delivery_persons WHERE id = ?', (person_id,))
    db.commit()
    log_activity(current_user.id, 'delete_delivery_person', f'Deleted delivery person #{person_id}: {person["name"] if person else "Unknown"}')
    db.close()
    flash('Delivery person deleted.', 'success')
    return redirect(url_for('delivery_persons'))


# ── Vehicle Types Management ─────────────────────────────────
@app.route('/vehicle-types')
@admin_required
def vehicle_types():
    db = get_db()
    vtypes = db.execute('SELECT * FROM vehicle_types ORDER BY name').fetchall()
    db.close()
    return render_template('vehicle_types.html', vehicle_types=vtypes)


@app.route('/vehicle-types/add', methods=['POST'])
@admin_required
def add_vehicle_type():
    name = request.form.get('name', '').strip()
    if not name:
        flash('Vehicle type name is required.', 'danger')
        return redirect(url_for('vehicle_types'))
    db = get_db()
    existing = db.execute('SELECT id FROM vehicle_types WHERE name = ?', (name,)).fetchone()
    if existing:
        flash('A vehicle type with that name already exists.', 'warning')
        db.close()
        return redirect(url_for('vehicle_types'))
    db.execute('INSERT INTO vehicle_types (name) VALUES (?)', (name,))
    db.commit()
    log_activity(current_user.id, 'add_vehicle_type', f'Created vehicle type: {name}')
    db.close()
    flash(f'Vehicle type "{name}" created.', 'success')
    return redirect(url_for('vehicle_types'))


@app.route('/vehicle-types/delete/<int:vtype_id>')
@admin_required
def delete_vehicle_type(vtype_id):
    db = get_db()
    vt = db.execute('SELECT name FROM vehicle_types WHERE id = ?', (vtype_id,)).fetchone()
    db.execute('DELETE FROM vehicle_types WHERE id = ?', (vtype_id,))
    db.commit()
    log_activity(current_user.id, 'delete_vehicle_type', f'Deleted vehicle type: {vt["name"] if vt else "Unknown"}')
    db.close()
    flash('Vehicle type deleted.', 'success')
    return redirect(url_for('vehicle_types'))


# ── Activity Logs (Admin) ────────────────────────────────────
@app.route('/logs')
@admin_required
def activity_logs():
    db = get_db()
    user_filter = request.args.get('user', '')
    action_filter = request.args.get('action', '')

    query = '''
        SELECT l.*, u.full_name, u.username
        FROM activity_logs l
        LEFT JOIN users u ON l.user_id = u.id
        WHERE 1=1
    '''
    params = []

    if user_filter:
        query += ' AND l.user_id = ?'
        params.append(int(user_filter))

    if action_filter:
        query += ' AND l.action = ?'
        params.append(action_filter)

    query += ' ORDER BY l.created_at DESC LIMIT 500'

    logs = db.execute(query, params).fetchall()
    users = db.execute("SELECT id, full_name FROM users ORDER BY full_name").fetchall()
    db.close()

    return render_template('logs.html', logs=logs, users=users,
                           user_filter=user_filter, action_filter=action_filter)


# ── Messaging ────────────────────────────────────────────────
@app.route('/messages')
@login_required
def messages_inbox():
    db = get_db()
    # Get all users this person has conversations with
    conversations = db.execute('''
        SELECT u.id, u.full_name, u.profile_pic, u.role,
            (SELECT COUNT(*) FROM messages WHERE sender_id = u.id AND receiver_id = ? AND is_read = 0) as unread,
            (SELECT MAX(created_at) FROM messages WHERE (sender_id = u.id AND receiver_id = ?) OR (sender_id = ? AND receiver_id = u.id)) as last_msg_time,
            (SELECT message FROM messages WHERE ((sender_id = u.id AND receiver_id = ?) OR (sender_id = ? AND receiver_id = u.id)) ORDER BY created_at DESC LIMIT 1) as last_msg
        FROM users u
        WHERE u.id != ? AND u.is_active = 1
        ORDER BY last_msg_time DESC NULLS LAST, u.full_name ASC
    ''', (current_user.id, current_user.id, current_user.id, current_user.id, current_user.id, current_user.id)).fetchall()
    db.close()
    return render_template('messages_inbox.html', conversations=conversations)


@app.route('/messages/<int:user_id>')
@login_required
def messages_chat(user_id):
    db = get_db()
    other_user = db.execute('SELECT id, full_name, profile_pic, role FROM users WHERE id = ?', (user_id,)).fetchone()
    if not other_user:
        db.close()
        flash('User not found.', 'danger')
        return redirect(url_for('messages_inbox'))

    # Mark as read
    db.execute('UPDATE messages SET is_read = 1 WHERE sender_id = ? AND receiver_id = ? AND is_read = 0',
               (user_id, current_user.id))
    db.commit()

    chat = db.execute('''
        SELECT m.*, s.full_name as sender_name, s.profile_pic as sender_pic
        FROM messages m
        LEFT JOIN users s ON m.sender_id = s.id
        WHERE (m.sender_id = ? AND m.receiver_id = ?) OR (m.sender_id = ? AND m.receiver_id = ?)
        ORDER BY m.created_at ASC
    ''', (current_user.id, user_id, user_id, current_user.id)).fetchall()
    db.close()
    return render_template('messages_chat.html', chat=chat, other_user=other_user)


@app.route('/messages/<int:user_id>/send', methods=['POST'])
@login_required
def send_message(user_id):
    message_text = request.form.get('message', '').strip()
    attachment_file = request.files.get('attachment')
    attachment_filename = None

    if attachment_file and attachment_file.filename:
        allowed_ext = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'zip'}
        ext = attachment_file.filename.rsplit('.', 1)[1].lower() if '.' in attachment_file.filename else ''
        if ext in allowed_ext:
            safe_name = secure_filename(attachment_file.filename)
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            attachment_filename = f'msg_{current_user.id}_{timestamp}_{safe_name}'
            attachment_file.save(os.path.join(UPLOAD_FOLDER, attachment_filename))
        else:
            flash('Invalid attachment type.', 'danger')
            return redirect(url_for('messages_chat', user_id=user_id))

    if not message_text and not attachment_filename:
        flash('Please enter a message or attach a file.', 'warning')
        return redirect(url_for('messages_chat', user_id=user_id))

    db = get_db()
    db.execute('INSERT INTO messages (sender_id, receiver_id, message, attachment) VALUES (?, ?, ?, ?)',
               (current_user.id, user_id, message_text, attachment_filename))
    db.commit()
    db.close()

    # Send notification
    preview = (message_text[:50] + '...') if len(message_text) > 50 else message_text
    if attachment_filename and not message_text:
        preview = 'Sent an attachment'
    send_notification(user_id, 'New Message', f'{current_user.full_name}: {preview}', 'message', url_for('messages_chat', user_id=current_user.id))

    return redirect(url_for('messages_chat', user_id=user_id))


@app.route('/messages/<int:user_id>/new')
@login_required
def get_new_messages(user_id):
    after_id = request.args.get('after', 0, type=int)
    db = get_db()
    # Mark incoming as read
    db.execute('UPDATE messages SET is_read = 1 WHERE sender_id = ? AND receiver_id = ? AND is_read = 0',
               (user_id, current_user.id))
    db.commit()
    msgs = db.execute('''
        SELECT m.id, m.sender_id, m.message, m.attachment, m.created_at, s.full_name as sender_name
        FROM messages m
        LEFT JOIN users s ON m.sender_id = s.id
        WHERE m.id > ? AND ((m.sender_id = ? AND m.receiver_id = ?) OR (m.sender_id = ? AND m.receiver_id = ?))
        ORDER BY m.created_at ASC
    ''', (after_id, current_user.id, user_id, user_id, current_user.id)).fetchall()
    db.close()
    result = []
    for m in msgs:
        result.append({
            'id': m['id'],
            'sender_id': m['sender_id'],
            'message': m['message'] or '',
            'attachment': m['attachment'] or '',
            'created_at': m['created_at'] or '',
            'sender_name': m['sender_name'] or ''
        })
    return jsonify(result)


@app.route('/messages/<int:user_id>/send-ajax', methods=['POST'])
@login_required
def send_message_ajax(user_id):
    message_text = request.form.get('message', '').strip()
    attachment_file = request.files.get('attachment')
    attachment_filename = None

    if attachment_file and attachment_file.filename:
        allowed_ext = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'zip'}
        ext = attachment_file.filename.rsplit('.', 1)[1].lower() if '.' in attachment_file.filename else ''
        if ext in allowed_ext:
            safe_name = secure_filename(attachment_file.filename)
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            attachment_filename = f'msg_{current_user.id}_{timestamp}_{safe_name}'
            attachment_file.save(os.path.join(UPLOAD_FOLDER, attachment_filename))
        else:
            return jsonify({'ok': False, 'error': 'Invalid file type'}), 400

    if not message_text and not attachment_filename:
        return jsonify({'ok': False, 'error': 'Empty message'}), 400

    db = get_db()
    db.execute('INSERT INTO messages (sender_id, receiver_id, message, attachment) VALUES (?, ?, ?, ?)',
               (current_user.id, user_id, message_text, attachment_filename))
    db.commit()
    db.close()

    preview = (message_text[:50] + '...') if len(message_text) > 50 else message_text
    if attachment_filename and not message_text:
        preview = 'Sent an attachment'
    send_notification(user_id, 'New Message', f'{current_user.full_name}: {preview}', 'message', url_for('messages_chat', user_id=current_user.id))

    return jsonify({'ok': True})


@app.route('/messages/unread-count')
@login_required
def unread_messages_count():
    db = get_db()
    count = db.execute('SELECT COUNT(*) FROM messages WHERE receiver_id = ? AND is_read = 0',
                      (current_user.id,)).fetchone()[0]
    db.close()
    return jsonify({'count': count})


@app.route('/messages/<int:user_id>/delete/<int:msg_id>', methods=['POST'])
@login_required
def delete_message(user_id, msg_id):
    db = get_db()
    msg = db.execute('SELECT * FROM messages WHERE id = ?', (msg_id,)).fetchone()
    if not msg or msg['sender_id'] != current_user.id:
        db.close()
        return jsonify({'ok': False, 'error': 'Cannot delete this message'}), 403
    # Delete attachment file if exists
    if msg['attachment']:
        att_path = os.path.join(UPLOAD_FOLDER, msg['attachment'])
        if os.path.exists(att_path):
            os.remove(att_path)
    db.execute('DELETE FROM messages WHERE id = ?', (msg_id,))
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/messages/<int:user_id>/clear', methods=['POST'])
@login_required
def clear_messages(user_id):
    db = get_db()
    # Get attachments to delete
    attachments = db.execute('''
        SELECT attachment FROM messages
        WHERE sender_id = ? AND receiver_id = ? AND attachment IS NOT NULL AND attachment != ''
    ''', (current_user.id, user_id)).fetchall()
    for a in attachments:
        att_path = os.path.join(UPLOAD_FOLDER, a['attachment'])
        if os.path.exists(att_path):
            os.remove(att_path)
    # Delete only messages sent by current user
    db.execute('DELETE FROM messages WHERE sender_id = ? AND receiver_id = ?',
               (current_user.id, user_id))
    db.commit()
    db.close()
    flash('Your messages have been cleared.', 'success')
    return redirect(url_for('messages_chat', user_id=user_id))


# ── HR Module ─────────────────────────────────────────────────
@app.route('/hr/employees')
@login_required
def hr_employees():
    db = get_db()
    employees = db.execute('''
        SELECT e.*, u.full_name as linked_user_name
        FROM hr_employees e
        LEFT JOIN users u ON e.user_id = u.id
        ORDER BY e.employee_name
    ''').fetchall()
    db.close()
    return render_template('hr_employees.html', employees=employees)


@app.route('/hr/employees/add', methods=['GET', 'POST'])
@login_required
def add_hr_employee():
    db = get_db()
    if request.method == 'POST':
        employee_name = request.form.get('employee_name', '').strip()
        employee_id = request.form.get('employee_id', '').strip()
        phone = request.form.get('phone', '').strip()
        designation = request.form.get('designation', '').strip()
        joining_date = request.form.get('joining_date', '').strip()
        basic_salary = float(request.form.get('basic_salary') or 0)
        housing_allowance = float(request.form.get('housing_allowance') or 0)
        transport_allowance = float(request.form.get('transport_allowance') or 0)
        other_allowance = float(request.form.get('other_allowance') or 0)
        ticket_amount = float(request.form.get('ticket_amount') or 0)
        user_id = request.form.get('user_id') or None

        if not employee_name or not joining_date:
            flash('Employee name and joining date are required.', 'danger')
        else:
            db.execute('''
                INSERT INTO hr_employees (employee_name, employee_id, phone, designation, joining_date,
                    basic_salary, housing_allowance, transport_allowance, other_allowance, ticket_amount,
                    user_id, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (employee_name, employee_id, phone, designation, joining_date,
                  basic_salary, housing_allowance, transport_allowance, other_allowance, ticket_amount,
                  int(user_id) if user_id else None, current_user.id))
            db.commit()
            log_activity(current_user.id, 'add_hr_employee', f'Added employee: {employee_name}')
            db.close()
            flash('Employee added successfully.', 'success')
            return redirect(url_for('hr_employees'))

    users = db.execute("SELECT id, full_name FROM users WHERE is_active = 1").fetchall()
    db.close()
    return render_template('hr_employee_form.html', employee=None, users=users)


@app.route('/hr/employees/edit/<int:emp_id>', methods=['GET', 'POST'])
@login_required
def edit_hr_employee(emp_id):
    db = get_db()
    employee = db.execute('SELECT * FROM hr_employees WHERE id = ?', (emp_id,)).fetchone()
    if not employee:
        db.close()
        flash('Employee not found.', 'danger')
        return redirect(url_for('hr_employees'))

    if request.method == 'POST':
        employee_name = request.form.get('employee_name', '').strip()
        employee_id = request.form.get('employee_id', '').strip()
        phone = request.form.get('phone', '').strip()
        designation = request.form.get('designation', '').strip()
        joining_date = request.form.get('joining_date', '').strip()
        basic_salary = float(request.form.get('basic_salary') or 0)
        housing_allowance = float(request.form.get('housing_allowance') or 0)
        transport_allowance = float(request.form.get('transport_allowance') or 0)
        other_allowance = float(request.form.get('other_allowance') or 0)
        ticket_amount = float(request.form.get('ticket_amount') or 0)
        user_id = request.form.get('user_id') or None
        is_active = 1 if request.form.get('is_active') else 0

        if not employee_name or not joining_date:
            flash('Employee name and joining date are required.', 'danger')
        else:
            db.execute('''
                UPDATE hr_employees SET employee_name=?, employee_id=?, phone=?, designation=?, joining_date=?,
                    basic_salary=?, housing_allowance=?, transport_allowance=?, other_allowance=?, ticket_amount=?,
                    user_id=?, is_active=?
                WHERE id=?
            ''', (employee_name, employee_id, phone, designation, joining_date,
                  basic_salary, housing_allowance, transport_allowance, other_allowance, ticket_amount,
                  int(user_id) if user_id else None, is_active, emp_id))
            db.commit()
            log_activity(current_user.id, 'edit_hr_employee', f'Edited employee #{emp_id}: {employee_name}')
            db.close()
            flash('Employee updated successfully.', 'success')
            return redirect(url_for('hr_employees'))

    users = db.execute("SELECT id, full_name FROM users WHERE is_active = 1").fetchall()
    db.close()
    return render_template('hr_employee_form.html', employee=employee, users=users)


@app.route('/hr/employees/delete/<int:emp_id>')
@admin_required
def delete_hr_employee(emp_id):
    db = get_db()
    emp = db.execute('SELECT employee_name FROM hr_employees WHERE id = ?', (emp_id,)).fetchone()
    db.execute('DELETE FROM hr_leaves WHERE employee_id = ?', (emp_id,))
    db.execute('DELETE FROM hr_employees WHERE id = ?', (emp_id,))
    db.commit()
    log_activity(current_user.id, 'delete_hr_employee', f'Deleted employee #{emp_id}: {emp["employee_name"] if emp else "Unknown"}')
    db.close()
    flash('Employee deleted.', 'success')
    return redirect(url_for('hr_employees'))


@app.route('/hr/leaves')
@login_required
def hr_leaves():
    db = get_db()
    emp_filter = request.args.get('employee', '')
    status_filter = request.args.get('status', '')

    query = '''
        SELECT l.*, e.employee_name, e.employee_id as emp_code, e.joining_date
        FROM hr_leaves l
        LEFT JOIN hr_employees e ON l.employee_id = e.id
        WHERE 1=1
    '''
    params = []
    if emp_filter:
        query += ' AND l.employee_id = ?'
        params.append(int(emp_filter))
    if status_filter:
        query += ' AND l.status = ?'
        params.append(status_filter)
    query += ' ORDER BY l.created_at DESC'

    leaves = db.execute(query, params).fetchall()
    employees = db.execute("SELECT id, employee_name FROM hr_employees WHERE is_active = 1 ORDER BY employee_name").fetchall()
    users = db.execute("SELECT id, full_name FROM users WHERE is_active = 1 ORDER BY full_name").fetchall()
    db.close()
    return render_template('hr_leaves.html', leaves=leaves, employees=employees, users=users,
                           emp_filter=emp_filter, status_filter=status_filter)


@app.route('/hr/leaves/add', methods=['GET', 'POST'])
@login_required
def add_hr_leave():
    db = get_db()
    if request.method == 'POST':
        employee_id = request.form.get('employee_id')
        leave_start_date = request.form.get('leave_start_date', '').strip()
        leave_end_date = request.form.get('leave_end_date', '').strip()
        rejoin_date = request.form.get('rejoin_date', '').strip() or None
        leave_type = request.form.get('leave_type', 'annual')
        notes = request.form.get('notes', '').strip()
        status = request.form.get('status', 'pending')
        calc_mode = request.form.get('calc_mode', 'auto')

        if not employee_id or not leave_start_date or not leave_end_date:
            flash('Employee, vacation start and end dates are required.', 'danger')
        else:
            emp = db.execute('SELECT * FROM hr_employees WHERE id = ?', (int(employee_id),)).fetchone()
            if not emp:
                flash('Employee not found.', 'danger')
            else:
                start_dt = datetime.strptime(leave_start_date, '%Y-%m-%d').date()
                end_dt = datetime.strptime(leave_end_date, '%Y-%m-%d').date()
                joining_dt = datetime.strptime(emp['joining_date'], '%Y-%m-%d').date()

                vacation_days = (end_dt - start_dt).days + 1

                if calc_mode == 'manual':
                    working_days = int(request.form.get('working_days', 0))
                    daily_salary = float(request.form.get('daily_salary', 0))
                    leave_salary = float(request.form.get('leave_salary', 0))
                    travel_allowance = float(request.form.get('ticket_amount', 0))
                    total_amount = float(request.form.get('total_amount', 0))
                else:
                    # Get formula settings
                    settings = dict(db.execute('SELECT key, value FROM settings').fetchall())
                    divisor = int(settings.get('leave_salary_divisor', '365'))

                    # Build salary components based on settings
                    total_monthly = 0
                    if settings.get('leave_include_basic', '1') == '1':
                        total_monthly += emp['basic_salary']
                    if settings.get('leave_include_housing', '1') == '1':
                        total_monthly += emp['housing_allowance']
                    if settings.get('leave_include_transport', '1') == '1':
                        total_monthly += emp['transport_allowance']
                    if settings.get('leave_include_other', '1') == '1':
                        total_monthly += emp['other_allowance']

                    service_days = (start_dt - joining_dt).days + 1
                    if service_days < 0:
                        service_days = 0
                    working_days = service_days

                    daily_salary = round(total_monthly / divisor, 2)
                    leave_salary = round(daily_salary * service_days, 2)

                    if settings.get('leave_include_ticket', '1') == '1' and emp['ticket_amount']:
                        travel_allowance = round((emp['ticket_amount'] / divisor) * service_days, 2)
                    else:
                        travel_allowance = 0

                    total_amount = round(leave_salary + travel_allowance, 2)

                db.execute('''
                    INSERT INTO hr_leaves (employee_id, leave_start_date, leave_end_date, rejoin_date,
                        leave_type, vacation_days, working_days, daily_salary, leave_salary,
                        ticket_included, ticket_amount, total_amount, status, notes, calc_mode, created_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (int(employee_id), leave_start_date, leave_end_date, rejoin_date,
                      leave_type, vacation_days, working_days, daily_salary, leave_salary,
                      1, travel_allowance, total_amount, status, notes, calc_mode, current_user.id))
                db.commit()
                log_activity(current_user.id, 'add_hr_leave', f'Added vacation for {emp["employee_name"]}: {leave_start_date} to {leave_end_date}')
                db.close()
                flash(f'Vacation added ({calc_mode}). Attendance: {working_days} days, Total: {total_amount}', 'success')
                return redirect(url_for('hr_leaves'))

    employees = db.execute("SELECT * FROM hr_employees WHERE is_active = 1 ORDER BY employee_name").fetchall()
    settings = dict(db.execute('SELECT key, value FROM settings').fetchall())
    db.close()
    return render_template('hr_leave_form.html', leave=None, employees=employees, leave_settings=settings)


@app.route('/hr/leaves/edit/<int:leave_id>', methods=['GET', 'POST'])
@login_required
def edit_hr_leave(leave_id):
    db = get_db()
    leave = db.execute('SELECT * FROM hr_leaves WHERE id = ?', (leave_id,)).fetchone()
    if not leave:
        db.close()
        flash('Leave record not found.', 'danger')
        return redirect(url_for('hr_leaves'))

    if request.method == 'POST':
        employee_id = request.form.get('employee_id')
        leave_start_date = request.form.get('leave_start_date', '').strip()
        leave_end_date = request.form.get('leave_end_date', '').strip()
        rejoin_date = request.form.get('rejoin_date', '').strip() or None
        leave_type = request.form.get('leave_type', 'annual')
        notes = request.form.get('notes', '').strip()
        status = request.form.get('status', 'pending')
        calc_mode = request.form.get('calc_mode', 'auto')

        if not employee_id or not leave_start_date or not leave_end_date:
            flash('Employee, vacation start and end dates are required.', 'danger')
        else:
            emp = db.execute('SELECT * FROM hr_employees WHERE id = ?', (int(employee_id),)).fetchone()
            if not emp:
                flash('Employee not found.', 'danger')
            else:
                start_dt = datetime.strptime(leave_start_date, '%Y-%m-%d').date()
                end_dt = datetime.strptime(leave_end_date, '%Y-%m-%d').date()
                joining_dt = datetime.strptime(emp['joining_date'], '%Y-%m-%d').date()

                vacation_days = (end_dt - start_dt).days + 1

                if calc_mode == 'manual':
                    working_days = int(request.form.get('working_days', 0))
                    daily_salary = float(request.form.get('daily_salary', 0))
                    leave_salary = float(request.form.get('leave_salary', 0))
                    travel_allowance = float(request.form.get('ticket_amount', 0))
                    total_amount = float(request.form.get('total_amount', 0))
                else:
                    settings_dict = dict(db.execute('SELECT key, value FROM settings').fetchall())
                    divisor = int(settings_dict.get('leave_salary_divisor', '365'))

                    total_monthly = 0
                    if settings_dict.get('leave_include_basic', '1') == '1':
                        total_monthly += emp['basic_salary']
                    if settings_dict.get('leave_include_housing', '1') == '1':
                        total_monthly += emp['housing_allowance']
                    if settings_dict.get('leave_include_transport', '1') == '1':
                        total_monthly += emp['transport_allowance']
                    if settings_dict.get('leave_include_other', '1') == '1':
                        total_monthly += emp['other_allowance']

                    service_days = (start_dt - joining_dt).days + 1
                    if service_days < 0:
                        service_days = 0
                    working_days = service_days

                    daily_salary = round(total_monthly / divisor, 2)
                    leave_salary = round(daily_salary * service_days, 2)

                    if settings_dict.get('leave_include_ticket', '1') == '1' and emp['ticket_amount']:
                        travel_allowance = round((emp['ticket_amount'] / divisor) * service_days, 2)
                    else:
                        travel_allowance = 0

                    total_amount = round(leave_salary + travel_allowance, 2)

                db.execute('''
                    UPDATE hr_leaves SET employee_id=?, leave_start_date=?, leave_end_date=?, rejoin_date=?,
                        leave_type=?, vacation_days=?, working_days=?, daily_salary=?, leave_salary=?,
                        ticket_included=?, ticket_amount=?, total_amount=?, status=?, notes=?, calc_mode=?
                    WHERE id=?
                ''', (int(employee_id), leave_start_date, leave_end_date, rejoin_date,
                      leave_type, vacation_days, working_days, daily_salary, leave_salary,
                      1, travel_allowance, total_amount, status, notes, calc_mode, leave_id))
                db.commit()
                log_activity(current_user.id, 'edit_hr_leave', f'Edited vacation #{leave_id} for {emp["employee_name"]}')
                db.close()
                flash(f'Vacation updated ({calc_mode}). Attendance: {working_days} days, Total: {total_amount}', 'success')
                return redirect(url_for('hr_leaves'))

    employees = db.execute("SELECT * FROM hr_employees WHERE is_active = 1 ORDER BY employee_name").fetchall()
    settings = dict(db.execute('SELECT key, value FROM settings').fetchall())
    db.close()
    return render_template('hr_leave_form.html', leave=leave, employees=employees, leave_settings=settings)


@app.route('/hr/leaves/delete/<int:leave_id>')
@admin_required
def delete_hr_leave(leave_id):
    db = get_db()
    db.execute('DELETE FROM hr_leaves WHERE id = ?', (leave_id,))
    db.commit()
    log_activity(current_user.id, 'delete_hr_leave', f'Deleted leave #{leave_id}')
    db.close()
    flash('Leave record deleted.', 'success')
    return redirect(url_for('hr_leaves'))


@app.route('/hr/leaves/pay/<int:leave_id>', methods=['POST'])
@login_required
def pay_hr_leave(leave_id):
    db = get_db()
    leave = db.execute('SELECT * FROM hr_leaves WHERE id = ?', (leave_id,)).fetchone()
    if not leave:
        db.close()
        flash('Leave record not found.', 'danger')
        return redirect(url_for('hr_leaves'))

    amount = float(request.form.get('amount', 0))
    payment_method = request.form.get('payment_method', '').strip()
    payment_reference = request.form.get('payment_reference', '').strip()
    payment_date = request.form.get('payment_date', '').strip() or date.today().isoformat()
    paid_by = request.form.get('paid_by', '').strip()

    new_paid = (leave['amount_paid'] or 0) + amount
    new_status = leave['status']
    if new_paid >= (leave['total_amount'] or 0):
        new_status = 'paid'

    db.execute('''
        UPDATE hr_leaves SET amount_paid = ?, payment_method = ?, payment_reference = ?,
            payment_date = ?, paid_by = ?, status = ?
        WHERE id = ?
    ''', (new_paid, payment_method, payment_reference, payment_date, paid_by, new_status, leave_id))
    db.commit()

    emp = db.execute('SELECT employee_name FROM hr_employees WHERE id = ?', (leave['employee_id'],)).fetchone()
    emp_name = emp['employee_name'] if emp else 'Unknown'
    log_activity(current_user.id, 'pay_hr_leave', f'Payment {amount} for leave #{leave_id} ({emp_name})')
    db.close()
    flash(f'Payment of {amount:.2f} SAR recorded for {emp_name}.', 'success')
    return redirect(url_for('hr_leaves'))


@app.route('/settings/leave-formula', methods=['GET', 'POST'])
@admin_required
def settings_leave_formula():
    db = get_db()
    if request.method == 'POST':
        keys = ['leave_salary_divisor', 'leave_include_basic', 'leave_include_housing',
                'leave_include_transport', 'leave_include_other', 'leave_include_ticket', 'leave_calc_mode']
        for key in keys:
            if key in ['leave_salary_divisor', 'leave_calc_mode']:
                val = request.form.get(key, '')
            else:
                val = '1' if request.form.get(key) else '0'
            db.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', (key, val))
        db.commit()
        log_activity(current_user.id, 'update_settings', 'Updated leave salary formula settings')
        flash('Leave salary formula settings updated.', 'success')
        return redirect(url_for('settings_leave_formula'))

    settings = dict(db.execute('SELECT key, value FROM settings').fetchall())
    db.close()
    return render_template('settings_leave_formula.html', settings=settings)


@app.route('/hr/calculate', methods=['POST'])
@login_required
def hr_calculate_leave():
    """AJAX endpoint for auto-calculating vacation salary."""
    employee_id = request.form.get('employee_id')
    leave_start_date = request.form.get('leave_start_date', '').strip()
    leave_end_date = request.form.get('leave_end_date', '').strip()

    if not employee_id or not leave_start_date or not leave_end_date:
        return jsonify({'error': 'Missing required fields'}), 400

    db = get_db()
    emp = db.execute('SELECT * FROM hr_employees WHERE id = ?', (int(employee_id),)).fetchone()
    settings = dict(db.execute('SELECT key, value FROM settings').fetchall())
    db.close()

    if not emp:
        return jsonify({'error': 'Employee not found'}), 404

    try:
        start_dt = datetime.strptime(leave_start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(leave_end_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    joining_dt = datetime.strptime(emp['joining_date'], '%Y-%m-%d').date()
    divisor = int(settings.get('leave_salary_divisor', '365'))

    # Build salary based on settings
    total_monthly = 0
    if settings.get('leave_include_basic', '1') == '1':
        total_monthly += emp['basic_salary']
    if settings.get('leave_include_housing', '1') == '1':
        total_monthly += emp['housing_allowance']
    if settings.get('leave_include_transport', '1') == '1':
        total_monthly += emp['transport_allowance']
    if settings.get('leave_include_other', '1') == '1':
        total_monthly += emp['other_allowance']

    service_days = (start_dt - joining_dt).days + 1
    if service_days < 0:
        service_days = 0

    vacation_days = (end_dt - start_dt).days + 1

    daily_salary = round(total_monthly / divisor, 2)
    years_employed = round(service_days / 365, 2)

    leave_salary = round(daily_salary * service_days, 2)

    if settings.get('leave_include_ticket', '1') == '1' and emp['ticket_amount']:
        travel_allowance = round((emp['ticket_amount'] / divisor) * service_days, 2)
    else:
        travel_allowance = 0

    total_amount = round(leave_salary + travel_allowance, 2)

    return jsonify({
        'employee_name': emp['employee_name'],
        'joining_date': emp['joining_date'],
        'total_monthly_salary': total_monthly,
        'daily_salary': daily_salary,
        'divisor': divisor,
        'service_days': service_days,
        'vacation_days': vacation_days,
        'years_employed': years_employed,
        'leave_salary': leave_salary,
        'ticket_amount_annual': emp['ticket_amount'],
        'travel_allowance': travel_allowance,
        'total_amount': total_amount
    })


@app.route('/hr/leaves/export')
@login_required
def export_hr_leaves():
    db = get_db()
    leaves = db.execute('''
        SELECT l.*, e.employee_name, e.employee_id as emp_code, e.joining_date
        FROM hr_leaves l
        LEFT JOIN hr_employees e ON l.employee_id = e.id
        ORDER BY l.created_at DESC
    ''').fetchall()
    db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Vacation Records'
    ws.append(['#', 'Employee', 'Emp ID', 'Joining Date', 'Vacation Start', 'Vacation End', 'Rejoin Date',
               'Type', 'Vacation Days', 'Attendance Days', 'Daily Salary', 'Leave Salary',
               'Travel Allowance', 'Total Amount', 'Status', 'Notes'])
    for l in leaves:
        ws.append([l['id'], l['employee_name'], l['emp_code'] or '', l['joining_date'] or '',
                   l['leave_start_date'], l['leave_end_date'], l['rejoin_date'] or '',
                   l['leave_type'], l['vacation_days'], l['working_days'],
                   l['daily_salary'], l['leave_salary'],
                   l['ticket_amount'],
                   l['total_amount'], l['status'], l['notes'] or ''])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='vacation_records.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Vehicle Document Expiry Manager ──────────────────────────
@app.route('/vehicles')
@login_required
def vehicles():
    db = get_db()
    vehicle_list = db.execute('''
        SELECT v.*, dp.name as linked_driver_name, dp.mobile as driver_mobile,
            (SELECT COUNT(*) FROM vehicle_documents WHERE vehicle_id = v.id) as doc_count,
            (SELECT MIN(expiry_date) FROM vehicle_documents WHERE vehicle_id = v.id AND expiry_date >= date('now')) as next_expiry
        FROM vehicles v
        LEFT JOIN delivery_persons dp ON v.driver_id = dp.id
        ORDER BY v.vehicle_name
    ''').fetchall()
    db.close()
    today = date.today().isoformat()
    return render_template('vehicles.html', vehicles=vehicle_list, today=today)


@app.route('/vehicles/add', methods=['GET', 'POST'])
@login_required
def add_vehicle():
    db = get_db()
    if request.method == 'POST':
        vehicle_name = request.form.get('vehicle_name', '').strip()
        plate_number = request.form.get('plate_number', '').strip()
        vehicle_type = request.form.get('vehicle_type', '').strip()
        driver_name = request.form.get('driver_name', '').strip()
        driver_id = request.form.get('driver_id') or None

        if not vehicle_name or not plate_number:
            flash('Vehicle name and plate number are required.', 'danger')
        else:
            db.execute('''
                INSERT INTO vehicles (vehicle_name, plate_number, vehicle_type, driver_name, driver_id, created_by)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (vehicle_name, plate_number, vehicle_type, driver_name,
                  int(driver_id) if driver_id else None, current_user.id))
            db.commit()
            log_activity(current_user.id, 'add_vehicle', f'Added vehicle: {vehicle_name} ({plate_number})')
            db.close()
            flash('Vehicle added successfully.', 'success')
            return redirect(url_for('vehicles'))

    drivers = db.execute("SELECT id, name FROM delivery_persons WHERE is_active = 1 ORDER BY name").fetchall()
    db.close()
    return render_template('vehicle_form.html', vehicle=None, drivers=drivers)


@app.route('/vehicles/edit/<int:vehicle_id>', methods=['GET', 'POST'])
@login_required
def edit_vehicle(vehicle_id):
    db = get_db()
    vehicle = db.execute('SELECT * FROM vehicles WHERE id = ?', (vehicle_id,)).fetchone()
    if not vehicle:
        db.close()
        flash('Vehicle not found.', 'danger')
        return redirect(url_for('vehicles'))

    if request.method == 'POST':
        vehicle_name = request.form.get('vehicle_name', '').strip()
        plate_number = request.form.get('plate_number', '').strip()
        vehicle_type = request.form.get('vehicle_type', '').strip()
        driver_name = request.form.get('driver_name', '').strip()
        driver_id = request.form.get('driver_id') or None
        is_active = 1 if request.form.get('is_active') else 0

        if not vehicle_name or not plate_number:
            flash('Vehicle name and plate number are required.', 'danger')
        else:
            db.execute('''
                UPDATE vehicles SET vehicle_name=?, plate_number=?, vehicle_type=?, driver_name=?, driver_id=?, is_active=?
                WHERE id=?
            ''', (vehicle_name, plate_number, vehicle_type, driver_name,
                  int(driver_id) if driver_id else None, is_active, vehicle_id))
            db.commit()
            log_activity(current_user.id, 'edit_vehicle', f'Edited vehicle #{vehicle_id}: {vehicle_name}')
            db.close()
            flash('Vehicle updated successfully.', 'success')
            return redirect(url_for('vehicles'))

    drivers = db.execute("SELECT id, name FROM delivery_persons WHERE is_active = 1 ORDER BY name").fetchall()
    db.close()
    return render_template('vehicle_form.html', vehicle=vehicle, drivers=drivers)


@app.route('/vehicles/delete/<int:vehicle_id>')
@admin_required
def delete_vehicle(vehicle_id):
    db = get_db()
    v = db.execute('SELECT vehicle_name FROM vehicles WHERE id = ?', (vehicle_id,)).fetchone()
    db.execute('DELETE FROM vehicle_documents WHERE vehicle_id = ?', (vehicle_id,))
    db.execute('DELETE FROM vehicles WHERE id = ?', (vehicle_id,))
    db.commit()
    log_activity(current_user.id, 'delete_vehicle', f'Deleted vehicle #{vehicle_id}: {v["vehicle_name"] if v else "Unknown"}')
    db.close()
    flash('Vehicle deleted.', 'success')
    return redirect(url_for('vehicles'))


@app.route('/vehicles/<int:vehicle_id>/documents')
@login_required
def vehicle_documents(vehicle_id):
    db = get_db()
    vehicle = db.execute('''
        SELECT v.*, dp.name as driver_name_linked, dp.mobile as driver_mobile
        FROM vehicles v
        LEFT JOIN delivery_persons dp ON v.driver_id = dp.id
        WHERE v.id = ?
    ''', (vehicle_id,)).fetchone()
    if not vehicle:
        db.close()
        flash('Vehicle not found.', 'danger')
        return redirect(url_for('vehicles'))

    docs = db.execute('''
        SELECT * FROM vehicle_documents WHERE vehicle_id = ?
        ORDER BY expiry_date ASC
    ''', (vehicle_id,)).fetchall()
    db.close()
    today = date.today().isoformat()
    return render_template('vehicle_documents.html', vehicle=vehicle, documents=docs, today=today)


@app.route('/vehicles/<int:vehicle_id>/documents/add', methods=['GET', 'POST'])
@login_required
def add_vehicle_document(vehicle_id):
    db = get_db()
    vehicle = db.execute('SELECT * FROM vehicles WHERE id = ?', (vehicle_id,)).fetchone()
    if not vehicle:
        db.close()
        flash('Vehicle not found.', 'danger')
        return redirect(url_for('vehicles'))

    if request.method == 'POST':
        document_type = request.form.get('document_type', '').strip()
        document_number = request.form.get('document_number', '').strip()
        issue_date = request.form.get('issue_date') or None
        expiry_date = request.form.get('expiry_date', '').strip()
        reminder_days = int(request.form.get('reminder_days') or 30)
        notes = request.form.get('notes', '').strip()

        attachment_filename = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file and file.filename:
                allowed_ext = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf'}
                ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                if ext in allowed_ext:
                    safe_name = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                    attachment_filename = f'vdoc_{vehicle_id}_{timestamp}_{safe_name}'
                    file.save(os.path.join(UPLOAD_FOLDER, attachment_filename))

        if not document_type or not expiry_date:
            flash('Document type and expiry date are required.', 'danger')
        else:
            db.execute('''
                INSERT INTO vehicle_documents (vehicle_id, document_type, document_number, issue_date,
                    expiry_date, reminder_days, attachment, notes, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (vehicle_id, document_type, document_number, issue_date,
                  expiry_date, reminder_days, attachment_filename, notes, current_user.id))
            db.commit()
            log_activity(current_user.id, 'add_vehicle_document',
                         f'Added {document_type} for vehicle {vehicle["vehicle_name"]}')
            db.close()
            flash('Document added successfully.', 'success')
            return redirect(url_for('vehicle_documents', vehicle_id=vehicle_id))

    db.close()
    return render_template('vehicle_document_form.html', vehicle=vehicle, document=None)


@app.route('/vehicles/<int:vehicle_id>/documents/edit/<int:doc_id>', methods=['GET', 'POST'])
@login_required
def edit_vehicle_document(vehicle_id, doc_id):
    db = get_db()
    vehicle = db.execute('SELECT * FROM vehicles WHERE id = ?', (vehicle_id,)).fetchone()
    doc = db.execute('SELECT * FROM vehicle_documents WHERE id = ? AND vehicle_id = ?', (doc_id, vehicle_id)).fetchone()
    if not vehicle or not doc:
        db.close()
        flash('Document not found.', 'danger')
        return redirect(url_for('vehicles'))

    if request.method == 'POST':
        document_type = request.form.get('document_type', '').strip()
        document_number = request.form.get('document_number', '').strip()
        issue_date = request.form.get('issue_date') or None
        expiry_date = request.form.get('expiry_date', '').strip()
        reminder_days = int(request.form.get('reminder_days') or 30)
        notes = request.form.get('notes', '').strip()

        attachment_filename = doc['attachment']
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file and file.filename:
                allowed_ext = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf'}
                ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                if ext in allowed_ext:
                    safe_name = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                    attachment_filename = f'vdoc_{vehicle_id}_{timestamp}_{safe_name}'
                    file.save(os.path.join(UPLOAD_FOLDER, attachment_filename))

        if not document_type or not expiry_date:
            flash('Document type and expiry date are required.', 'danger')
        else:
            db.execute('''
                UPDATE vehicle_documents SET document_type=?, document_number=?, issue_date=?,
                    expiry_date=?, reminder_days=?, attachment=?, notes=?
                WHERE id=?
            ''', (document_type, document_number, issue_date,
                  expiry_date, reminder_days, attachment_filename, notes, doc_id))
            db.commit()
            log_activity(current_user.id, 'edit_vehicle_document',
                         f'Edited document #{doc_id} for vehicle {vehicle["vehicle_name"]}')
            db.close()
            flash('Document updated successfully.', 'success')
            return redirect(url_for('vehicle_documents', vehicle_id=vehicle_id))

    db.close()
    return render_template('vehicle_document_form.html', vehicle=vehicle, document=doc)


@app.route('/vehicles/<int:vehicle_id>/documents/delete/<int:doc_id>')
@admin_required
def delete_vehicle_document(vehicle_id, doc_id):
    db = get_db()
    doc = db.execute('SELECT * FROM vehicle_documents WHERE id = ? AND vehicle_id = ?', (doc_id, vehicle_id)).fetchone()
    if doc and doc['attachment']:
        att_path = os.path.join(UPLOAD_FOLDER, doc['attachment'])
        if os.path.exists(att_path):
            os.remove(att_path)
    db.execute('DELETE FROM vehicle_documents WHERE id = ?', (doc_id,))
    db.commit()
    log_activity(current_user.id, 'delete_vehicle_document', f'Deleted document #{doc_id}')
    db.close()
    flash('Document deleted.', 'success')
    return redirect(url_for('vehicle_documents', vehicle_id=vehicle_id))


@app.route('/vehicles/expiring')
@login_required
def vehicles_expiring():
    """Show all documents expiring within their reminder period or already expired."""
    db = get_db()
    docs = db.execute('''
        SELECT vd.*, v.vehicle_name, v.plate_number
        FROM vehicle_documents vd
        LEFT JOIN vehicles v ON vd.vehicle_id = v.id
        WHERE v.is_active = 1
        ORDER BY vd.expiry_date ASC
    ''').fetchall()
    db.close()
    today = date.today()
    expiring = []
    expired = []
    for d in docs:
        exp_date = datetime.strptime(d['expiry_date'], '%Y-%m-%d').date()
        reminder_date = exp_date - timedelta(days=d['reminder_days'])
        if exp_date < today:
            expired.append(d)
        elif reminder_date <= today:
            expiring.append(d)
    return render_template('vehicles_expiring.html', expiring=expiring, expired=expired, today=today.isoformat())


# ── Settings ─────────────────────────────────────────────────
@app.route('/settings')
@admin_required
def settings():
    return redirect(url_for('print_settings'))


@app.route('/settings/print', methods=['GET', 'POST'])
@admin_required
def print_settings():
    db = get_db()
    if request.method == 'POST':
        keys = ['print_paper_size', 'print_orientation', 'print_company_name', 'print_show_logo']
        for k in keys:
            val = request.form.get(k, '')
            if k == 'print_show_logo':
                val = '1' if request.form.get(k) else '0'
            db.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', (k, val))
        db.commit()
        log_activity(current_user.id, 'update_settings', 'Updated print settings')
        flash('Print settings saved.', 'success')
        db.close()
        return redirect(url_for('print_settings'))

    rows = db.execute('SELECT key, value FROM settings').fetchall()
    db.close()
    s = {r['key']: r['value'] for r in rows}
    return render_template('settings_print.html', settings=s)


@app.route('/settings/roles')
@admin_required
def role_settings():
    db = get_db()
    users_by_role = db.execute('''
        SELECT id, full_name, username, role, is_active FROM users ORDER BY
        CASE role WHEN 'super_admin' THEN 1 WHEN 'admin' THEN 2 ELSE 3 END, full_name
    ''').fetchall()
    db.close()
    return render_template('settings_roles.html', users=users_by_role)


@app.route('/settings/pin', methods=['GET', 'POST'])
@admin_required
def pin_settings():
    db = get_db()
    if request.method == 'POST':
        pin_keys = ['pin_on_add', 'pin_on_edit', 'pin_on_delete', 'pin_on_tasks', 'pin_on_deliveries', 'pin_on_hr', 'pin_on_vehicles']
        for k in pin_keys:
            val = '1' if request.form.get(k) else '0'
            db.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', (k, val))
        db.commit()
        log_activity(current_user.id, 'update_settings', 'Updated PIN settings')
        flash('PIN settings saved.', 'success')
        db.close()
        return redirect(url_for('pin_settings'))

    rows = db.execute('SELECT key, value FROM settings').fetchall()
    db.close()
    s = {r['key']: r['value'] for r in rows}
    return render_template('settings_pin.html', settings=s)


# ── Backup ───────────────────────────────────────────────────
backup_thread = None
backup_stop_event = threading.Event()


def do_backup():
    """Create a timestamped backup of the database."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f'backup_{timestamp}.db'
    src = DB_PATH
    dst = os.path.join(BACKUP_FOLDER, backup_name)
    shutil.copy2(src, dst)
    # Keep only the last 20 backups
    backups = sorted(
        [f for f in os.listdir(BACKUP_FOLDER) if f.startswith('backup_') and f.endswith('.db')],
        reverse=True
    )
    for old in backups[20:]:
        os.remove(os.path.join(BACKUP_FOLDER, old))
    return backup_name


def auto_backup_worker():
    """Background thread that backs up the DB at the configured interval."""
    while not backup_stop_event.is_set():
        try:
            db = get_db()
            rows = db.execute("SELECT key, value FROM settings WHERE key IN ('auto_backup', 'backup_interval_minutes')").fetchall()
            db.close()
            s = {r['key']: r['value'] for r in rows}
            if s.get('auto_backup') == '1':
                interval = int(s.get('backup_interval_minutes', '60'))
                do_backup()
                backup_stop_event.wait(interval * 60)
            else:
                backup_stop_event.wait(30)  # check again in 30s
        except Exception:
            backup_stop_event.wait(60)


def start_auto_backup():
    global backup_thread
    if backup_thread and backup_thread.is_alive():
        return
    backup_stop_event.clear()
    backup_thread = threading.Thread(target=auto_backup_worker, daemon=True)
    backup_thread.start()


@app.route('/settings/backup', methods=['GET', 'POST'])
@admin_required
def backup_settings():
    db = get_db()
    if request.method == 'POST':
        auto_backup = '1' if request.form.get('auto_backup') else '0'
        interval = request.form.get('backup_interval_minutes', '60')
        try:
            interval = str(max(1, int(interval)))
        except ValueError:
            interval = '60'
        db.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', ('auto_backup', auto_backup))
        db.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', ('backup_interval_minutes', interval))
        db.commit()
        log_activity(current_user.id, 'update_settings', f'Updated backup settings: auto={auto_backup}, interval={interval}min')
        flash('Backup settings saved.', 'success')
        db.close()
        # Restart the backup thread so it picks up new settings
        backup_stop_event.set()
        _time.sleep(0.2)
        start_auto_backup()
        return redirect(url_for('backup_settings'))

    rows = db.execute('SELECT key, value FROM settings').fetchall()
    s = {r['key']: r['value'] for r in rows}

    # List existing backups
    backups = []
    if os.path.exists(BACKUP_FOLDER):
        for f in sorted(os.listdir(BACKUP_FOLDER), reverse=True):
            if f.startswith('backup_') and f.endswith('.db'):
                fpath = os.path.join(BACKUP_FOLDER, f)
                size_mb = round(os.path.getsize(fpath) / (1024 * 1024), 2)
                backups.append({'name': f, 'size': size_mb,
                                'date': f.replace('backup_', '').replace('.db', '').replace('_', ' ')})
    db.close()
    return render_template('settings_backup.html', settings=s, backups=backups)


# ── Reports ──────────────────────────────────────────────────
@app.route('/reports/payments')
@login_required
def report_payments():
    db = get_db()
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    driver_filter = request.args.get('driver', '')
    method_filter = request.args.get('method', '')

    query = '''
        SELECT dp.*, d.delivery_note_number, d.customer_name, d.transportation_charge,
               d.delivery_person_id, drv.name as driver_name, u.full_name as recorded_by
        FROM delivery_payments dp
        JOIN deliveries d ON dp.delivery_id = d.id
        LEFT JOIN delivery_persons drv ON d.delivery_person_id = drv.id
        LEFT JOIN users u ON dp.created_by = u.id
        WHERE 1=1
    '''
    params = []
    if date_from:
        query += ' AND dp.paid_date >= ?'
        params.append(date_from)
    if date_to:
        query += ' AND dp.paid_date <= ?'
        params.append(date_to)
    if driver_filter:
        query += ' AND d.delivery_person_id = ?'
        params.append(int(driver_filter))
    if method_filter:
        query += ' AND dp.payment_method = ?'
        params.append(method_filter)
    query += ' ORDER BY dp.paid_date DESC, dp.created_at DESC'

    payments = db.execute(query, params).fetchall()
    drivers = db.execute("SELECT id, name FROM delivery_persons WHERE is_active = 1 ORDER BY name").fetchall()
    total_amount = sum(p['amount'] for p in payments)
    db.close()

    return render_template('report_payments.html', payments=payments, drivers=drivers,
                           date_from=date_from, date_to=date_to, driver_filter=driver_filter,
                           method_filter=method_filter, total_amount=total_amount)


@app.route('/reports/driver-ledger')
@login_required
def report_driver_ledger():
    db = get_db()
    driver_id = request.args.get('driver', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    drivers = db.execute("SELECT id, name FROM delivery_persons ORDER BY name").fetchall()
    ledger = []
    driver_info = None
    total_charge = 0
    total_paid = 0

    if driver_id:
        driver_info = db.execute("SELECT * FROM delivery_persons WHERE id = ?", (int(driver_id),)).fetchone()
        query = '''
            SELECT d.*, u.full_name as creator_name
            FROM deliveries d
            LEFT JOIN users u ON d.created_by = u.id
            WHERE d.delivery_person_id = ?
        '''
        params = [int(driver_id)]
        if date_from:
            query += ' AND d.delivery_date >= ?'
            params.append(date_from)
        if date_to:
            query += ' AND d.delivery_date <= ?'
            params.append(date_to)
        query += ' ORDER BY d.delivery_date DESC, d.created_at DESC'
        ledger = db.execute(query, params).fetchall()
        total_charge = sum(d['transportation_charge'] or 0 for d in ledger)
        total_paid = sum(d['amount_paid'] or 0 for d in ledger)

    db.close()
    return render_template('report_driver_ledger.html', drivers=drivers, ledger=ledger,
                           driver_info=driver_info, driver_filter=driver_id,
                           date_from=date_from, date_to=date_to,
                           total_charge=total_charge, total_paid=total_paid)


@app.route('/reports/delivery-accounts')
@login_required
def report_delivery_accounts():
    db = get_db()
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status_filter = request.args.get('status', '')

    query = '''
        SELECT d.*, dp.name as driver_name, u.full_name as creator_name
        FROM deliveries d
        LEFT JOIN delivery_persons dp ON d.delivery_person_id = dp.id
        LEFT JOIN users u ON d.created_by = u.id
        WHERE d.transportation_charge > 0
    '''
    params = []
    if date_from:
        query += ' AND d.delivery_date >= ?'
        params.append(date_from)
    if date_to:
        query += ' AND d.delivery_date <= ?'
        params.append(date_to)
    if status_filter == 'paid':
        query += ' AND d.charge_paid = 1'
    elif status_filter == 'unpaid':
        query += ' AND d.charge_paid = 0'
    elif status_filter == 'partial':
        query += ' AND d.charge_paid = 0 AND d.amount_paid > 0'
    query += ' ORDER BY d.delivery_date DESC, d.created_at DESC'

    accounts = db.execute(query, params).fetchall()
    total_charge = sum(d['transportation_charge'] or 0 for d in accounts)
    total_paid = sum(d['amount_paid'] or 0 for d in accounts)
    total_balance = total_charge - total_paid

    # Summary by driver
    driver_summary_q = '''
        SELECT dp.name as driver_name, dp.id as driver_id,
               COUNT(d.id) as delivery_count,
               SUM(d.transportation_charge) as total_charge,
               SUM(d.amount_paid) as total_paid,
               SUM(d.transportation_charge - COALESCE(d.amount_paid, 0)) as total_balance
        FROM deliveries d
        LEFT JOIN delivery_persons dp ON d.delivery_person_id = dp.id
        WHERE d.transportation_charge > 0
    '''
    ds_params = []
    if date_from:
        driver_summary_q += ' AND d.delivery_date >= ?'
        ds_params.append(date_from)
    if date_to:
        driver_summary_q += ' AND d.delivery_date <= ?'
        ds_params.append(date_to)
    driver_summary_q += ' GROUP BY dp.id ORDER BY total_balance DESC'
    driver_summary = db.execute(driver_summary_q, ds_params).fetchall()

    db.close()
    return render_template('report_delivery_accounts.html', accounts=accounts,
                           driver_summary=driver_summary,
                           date_from=date_from, date_to=date_to, status_filter=status_filter,
                           total_charge=total_charge, total_paid=total_paid, total_balance=total_balance)


@app.route('/reports/leave-salary')
@login_required
def report_leave_salary():
    db = get_db()
    employee_filter = request.args.get('employee', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status_filter = request.args.get('status', '')

    employees = db.execute('SELECT id, employee_name, employee_id FROM hr_employees ORDER BY employee_name').fetchall()

    query = '''
        SELECT l.*, e.employee_name, e.employee_id as emp_code, e.joining_date,
               e.basic_salary, e.housing_allowance, e.transport_allowance,
               e.other_allowance, e.ticket_amount as annual_ticket
        FROM hr_leaves l
        LEFT JOIN hr_employees e ON l.employee_id = e.id
        WHERE 1=1
    '''
    params = []
    if employee_filter:
        query += ' AND l.employee_id = ?'
        params.append(employee_filter)
    if date_from:
        query += ' AND l.leave_start_date >= ?'
        params.append(date_from)
    if date_to:
        query += ' AND l.leave_start_date <= ?'
        params.append(date_to)
    if status_filter:
        query += ' AND l.status = ?'
        params.append(status_filter)
    query += ' ORDER BY l.leave_start_date DESC'

    leaves = db.execute(query, params).fetchall()

    total_leave_salary = sum(l['leave_salary'] or 0 for l in leaves)
    total_ticket = sum(l['ticket_amount'] or 0 for l in leaves)
    total_amount = sum(l['total_amount'] or 0 for l in leaves)
    total_vacation_days = sum(l['vacation_days'] or 0 for l in leaves)
    total_working_days = sum(l['working_days'] or 0 for l in leaves)

    # Employee-wise summary
    emp_summary_q = '''
        SELECT e.id, e.employee_name, e.employee_id as emp_code,
               COUNT(l.id) as leave_count,
               SUM(l.vacation_days) as total_vacation_days,
               SUM(l.working_days) as total_working_days,
               SUM(l.leave_salary) as total_leave_salary,
               SUM(l.ticket_amount) as total_ticket,
               SUM(l.total_amount) as total_amount
        FROM hr_leaves l
        LEFT JOIN hr_employees e ON l.employee_id = e.id
        WHERE 1=1
    '''
    es_params = []
    if date_from:
        emp_summary_q += ' AND l.leave_start_date >= ?'
        es_params.append(date_from)
    if date_to:
        emp_summary_q += ' AND l.leave_start_date <= ?'
        es_params.append(date_to)
    if status_filter:
        emp_summary_q += ' AND l.status = ?'
        es_params.append(status_filter)
    emp_summary_q += ' GROUP BY e.id ORDER BY total_amount DESC'
    emp_summary = db.execute(emp_summary_q, es_params).fetchall()

    db.close()
    return render_template('report_leave_salary.html', leaves=leaves, employees=employees,
                           emp_summary=emp_summary,
                           employee_filter=employee_filter, date_from=date_from, date_to=date_to,
                           status_filter=status_filter,
                           total_leave_salary=total_leave_salary, total_ticket=total_ticket,
                           total_amount=total_amount, total_vacation_days=total_vacation_days,
                           total_working_days=total_working_days)


@app.route('/about')
@login_required
def about():
    return render_template('about.html')


@app.route('/backup/now', methods=['POST'])
@admin_required
def backup_now():
    try:
        name = do_backup()
        log_activity(current_user.id, 'manual_backup', f'Created manual backup: {name}')
        flash(f'Backup created: {name}', 'success')
    except Exception as e:
        flash(f'Backup failed: {str(e)}', 'danger')
    return redirect(url_for('backup_settings'))


@app.route('/backup/download/<filename>')
@admin_required
def download_backup(filename):
    safe_name = secure_filename(filename)
    fpath = os.path.join(BACKUP_FOLDER, safe_name)
    if not os.path.exists(fpath):
        flash('Backup not found.', 'danger')
        return redirect(url_for('backup_settings'))
    return send_file(fpath, download_name=safe_name, as_attachment=True)


@app.route('/backup/delete/<filename>', methods=['POST'])
@admin_required
def delete_backup(filename):
    safe_name = secure_filename(filename)
    fpath = os.path.join(BACKUP_FOLDER, safe_name)
    if os.path.exists(fpath):
        os.remove(fpath)
        flash('Backup deleted.', 'success')
    else:
        flash('Backup not found.', 'danger')
    return redirect(url_for('backup_settings'))


@app.route('/backup/import', methods=['POST'])
@admin_required
def import_backup():
    if 'backup_file' not in request.files:
        flash('No file selected.', 'danger')
        return redirect(url_for('backup_settings'))
    file = request.files['backup_file']
    if not file or not file.filename:
        flash('No file selected.', 'danger')
        return redirect(url_for('backup_settings'))
    if not file.filename.endswith('.db'):
        flash('Invalid file. Please upload a .db backup file.', 'danger')
        return redirect(url_for('backup_settings'))
    safe_name = secure_filename(file.filename)
    save_path = os.path.join(BACKUP_FOLDER, safe_name)
    file.save(save_path)
    log_activity(current_user.id, 'import_backup', f'Imported backup file: {safe_name}')
    flash(f'Backup imported: {safe_name}', 'success')
    return redirect(url_for('backup_settings'))


@app.route('/backup/restore/<filename>', methods=['POST'])
@admin_required
def restore_backup(filename):
    safe_name = secure_filename(filename)
    fpath = os.path.join(BACKUP_FOLDER, safe_name)
    if not os.path.exists(fpath):
        flash('Backup file not found.', 'danger')
        return redirect(url_for('backup_settings'))
    # Create a safety backup before restoring
    try:
        safety_name = 'pre_restore_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.db'
        shutil.copy2(DB_PATH, os.path.join(BACKUP_FOLDER, safety_name))
    except Exception:
        pass
    try:
        shutil.copy2(fpath, DB_PATH)
        log_activity(current_user.id, 'restore_backup', f'Restored backup: {safe_name}')
        flash(f'Database restored from {safe_name}. A safety backup was created before restore.', 'success')
    except Exception as e:
        flash(f'Restore failed: {str(e)}', 'danger')
    return redirect(url_for('backup_settings'))


@app.context_processor
def inject_settings():
    if current_user.is_authenticated:
        db = get_db()
        try:
            rows = db.execute('SELECT key, value FROM settings').fetchall()
            s = {r['key']: r['value'] for r in rows}
        except Exception:
            s = {}
        db.close()
        return {'app_settings': s, 'now': datetime.now}
    return {'app_settings': {}, 'now': datetime.now}


# ── Initialize ───────────────────────────────────────────────
def create_default_admin():
    db = get_db()
    existing = db.execute("SELECT id FROM users WHERE role = 'super_admin'").fetchone()
    if not existing:
        db.execute('''
            INSERT INTO users (username, password_hash, full_name, email, pin_hash, role)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', ('admin', _hash_pw('admin123'), 'Super Admin', 'admin@example.com', _hash_pw('1234'), 'super_admin'))
        db.commit()
        print("✅ Default super admin created: username='admin', password='admin123', PIN='1234'")
    db.close()


if __name__ == '__main__':
    init_db()
    create_default_admin()
    start_auto_backup()
    print("🚀 Server starting at http://127.0.0.1:5050")
    app.run(debug=True, host='0.0.0.0', port=5050)
